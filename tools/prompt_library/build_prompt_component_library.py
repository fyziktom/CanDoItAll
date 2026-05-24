from __future__ import annotations

import csv
import json
import re
import textwrap
import uuid
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
PACKS_ROOT = ROOT / "inputs" / "prompts packs"
DOCS_ROOT = ROOT / "docs" / "prompt-wizard"
OUTPUT_ROOT = ROOT / "output" / "prompt-library"
SPREADSHEET_ROOT = ROOT / "output" / "spreadsheet"
TEMPLATE_ROOT = OUTPUT_ROOT / "templates"

UUID_NAMESPACE = uuid.uuid5(uuid.NAMESPACE_URL, "https://candoitall.local/prompt-library/v1")

ALL_PROMPT_TYPES = [
    "architecture",
    "audit",
    "plan",
    "implementation",
    "refactor",
    "bugfix",
    "review",
    "testing",
    "validation",
    "performance",
    "security",
    "migration",
    "embedded",
    "ui",
]

ALL_PHASES = [
    "discovery",
    "architecture",
    "planning",
    "implementation",
    "verification",
    "delivery",
]


@dataclass(frozen=True)
class GroupDefinition:
    key: str
    name: str
    summary: str
    purpose: str
    ui_mode: str
    order: int


@dataclass
class ComponentDefinition:
    key: str
    name: str
    group: str
    block_kind: str
    summary: str
    template: str
    tags: list[str]
    prompt_types: list[str]
    blueprints: list[str]
    phases: list[str]
    stack_tags: list[str] = field(default_factory=list)
    toolbox_eligible: bool = False
    recommended: bool = False

    @property
    def id(self) -> str:
        return str(uuid.uuid5(UUID_NAMESPACE, f"component:{self.key}"))

    @property
    def template_tokens(self) -> list[str]:
        return sorted(set(re.findall(r"\{\{([^{}]+)\}\}", self.template)))

    @property
    def prompt_type_rules(self) -> str:
        return "|".join(self.prompt_types)

    @property
    def blueprint_rules(self) -> str:
        return "|".join(self.blueprints)

    @property
    def phase_rules(self) -> str:
        return "|".join(self.phases)

    def to_factory_seed(self) -> dict:
        return {
            "id": self.id,
            "key": self.key,
            "name": self.name,
            "blockKind": self.block_kind,
            "summary": self.summary,
            "content": self.template,
            "isRecommendedByDefault": self.recommended,
            "promptTypeRules": self.prompt_type_rules,
            "blueprintRules": self.blueprint_rules,
            "phaseRules": self.phase_rules,
            "group": self.group,
            "tags": self.tags,
            "stackTags": self.stack_tags,
            "toolboxEligible": self.toolbox_eligible,
            "templateTokens": self.template_tokens,
        }


@dataclass
class FlowAgentStep:
    order: int
    role_component_key: str
    blueprint_key: str
    phase: str
    goal: str
    block_keys: list[str]


@dataclass
class FlowDefinition:
    key: str
    name: str
    summary: str
    prompt_types: list[str]
    block_keys: list[str]
    agent_sequence: list[FlowAgentStep]

    @property
    def id(self) -> str:
        return str(uuid.uuid5(UUID_NAMESPACE, f"flow:{self.key}"))

    @property
    def prompt_type_rules(self) -> str:
        return "|".join(self.prompt_types)

    def to_factory_seed(self, block_lookup: dict[str, ComponentDefinition]) -> dict:
        block_ids = [block_lookup[key].id for key in self.block_keys]
        return {
            "id": self.id,
            "key": self.key,
            "name": self.name,
            "summary": self.summary,
            "blockIdsJson": json.dumps(block_ids, indent=2),
            "promptTypeRules": self.prompt_type_rules,
            "blockKeys": self.block_keys,
            "agentSequence": [
                {
                    "order": step.order,
                    "roleComponentId": block_lookup[step.role_component_key].id,
                    "roleComponentKey": step.role_component_key,
                    "blueprintKey": step.blueprint_key,
                    "phase": step.phase,
                    "goal": step.goal,
                    "blockKeys": step.block_keys,
                }
                for step in self.agent_sequence
            ],
        }


@dataclass
class BlueprintDefinition:
    key: str
    name: str
    prompt_type: str
    summary: str
    guidance: str
    recommended_flow_key: str
    recommended_block_keys: list[str]

    @property
    def id(self) -> str:
        return str(uuid.uuid5(UUID_NAMESPACE, f"blueprint:{self.key}"))

    def to_factory_seed(self, flows: dict[str, FlowDefinition]) -> dict:
        return {
            "id": self.id,
            "key": self.key,
            "name": self.name,
            "promptType": self.prompt_type,
            "summary": self.summary,
            "guidance": self.guidance,
            "recommendedFlowTemplateId": flows[self.recommended_flow_key].id,
            "recommendedFlowKey": self.recommended_flow_key,
            "recommendedBlockKeys": self.recommended_block_keys,
        }


@dataclass
class SimulationCase:
    key: str
    name: str
    summary: str
    flow_key: str
    stack_tags: list[str]
    required_groups: list[str]
    extra_block_keys: list[str]
    expected_roles: list[str]
    validation_focus: list[str]

    @property
    def id(self) -> str:
        return str(uuid.uuid5(UUID_NAMESPACE, f"simulation:{self.key}"))


def dedent(value: str) -> str:
    return textwrap.dedent(value).strip()


def stable_tags(*values: str) -> list[str]:
    return sorted({value for value in values if value})


def clean_markdown_doc(value: str) -> str:
    return dedent(value).replace("\n        ", "\n").strip()


GROUPS = [
    GroupDefinition("session-framing", "Session Framing and Role", "Defines the agent role, authority, and problem-solving posture.", "Use these blocks first so the model knows whether it is architecting, reviewing, planning, implementing, or validating.", "wizard-core", 1),
    GroupDefinition("mission-scope", "Mission, Scope, and Success", "Pins down the actual goal, boundaries, and end-state.", "These blocks stop prompt drift and make the session outcome measurable.", "wizard-core", 2),
    GroupDefinition("context-discovery", "Context Loading and Discovery", "Tells the agent what to read, inspect, and confirm before taking action.", "Most strong packs force the agent to read the repo, current state, and artifacts before proposing or changing anything.", "wizard-core", 3),
    GroupDefinition("guardrails", "Guardrails and Constraints", "Defines non-negotiables, limits, and safety rules.", "These blocks are the difference between a useful coding agent and an over-eager one.", "wizard-core", 4),
    GroupDefinition("workflow-orchestration", "Workflow Orchestration and Continuity", "Enforces phases, gates, status updates, and handoff continuity.", "The prompt packs consistently treat workflows as sequential, test-gated, and stateful.", "flow-core", 5),
    GroupDefinition("architecture-analysis", "Architecture and Analysis", "Reusable sections for architecture, audits, gap analysis, and design artifacts.", "These blocks are typically used by the first agent or by planning-focused sessions.", "flow-core", 6),
    GroupDefinition("planning-checklists", "Planning and Checklists", "Breaks work into milestones, files, tests, and acceptance gates.", "These blocks convert architecture into action without leaving the next agent to improvise.", "flow-core", 7),
    GroupDefinition("implementation-execution", "Implementation Execution", "Controls how code changes are made, sliced, and verified.", "The best packs force additive, low-risk implementation with continuous proof.", "flow-core", 8),
    GroupDefinition("validation-review", "Validation, Testing, and Review", "Makes quality evidence mandatory rather than optional.", "This group turns prompts into engineering workflows instead of writing exercises.", "validation-core", 9),
    GroupDefinition("output-handoff", "Output, Delivery, and Handoff", "Standardizes the final response, evidence, and next-step instructions.", "Strong packs require crisp handoff artifacts after each phase.", "wizard-core", 10),
    GroupDefinition("stack-profiles", "Stack Profiles", "Stack-specific constraints and guidance blocks for common technology areas.", "These are auto-applied or manually inserted based on the selected stack.", "stack-auto", 11),
    GroupDefinition("toolbox-snippets", "Toolbox Snippets", "Short optional inserts for concrete actions such as Docker tests or Playwright capture.", "These are the right-click or quick-add blocks that users can drop into a prompt.", "toolbox", 12),
]


def component(
    *,
    key: str,
    name: str,
    group: str,
    block_kind: str,
    summary: str,
    template: str,
    tags: Iterable[str],
    prompt_types: Iterable[str],
    blueprints: Iterable[str],
    phases: Iterable[str],
    stack_tags: Iterable[str] | None = None,
    toolbox_eligible: bool = False,
    recommended: bool = False,
) -> ComponentDefinition:
    return ComponentDefinition(
        key=key,
        name=name,
        group=group,
        block_kind=block_kind,
        summary=summary,
        template=dedent(template),
        tags=sorted(set(tags)),
        prompt_types=list(prompt_types),
        blueprints=list(blueprints),
        phases=list(phases),
        stack_tags=list(stack_tags or []),
        toolbox_eligible=toolbox_eligible,
        recommended=recommended,
    )


def build_role_components() -> list[ComponentDefinition]:
    role_specs = [
        {
            "key": "role-architecture-lead",
            "name": "Role: Architecture Lead",
            "summary": "Frames the session as architecture-first work with explicit boundaries, tradeoffs, and implementation-ready outputs.",
            "focus": [
                "produce an implementation-ready architecture for {{target_feature_or_problem}}",
                "make module boundaries, contracts, risks, and tradeoffs explicit",
                "avoid vague design language that cannot guide the next agent",
            ],
            "posture": [
                "inspect the current codebase and artifacts before proposing structural changes",
                "prefer the simplest architecture that still covers scale, quality, and maintenance needs",
                "tie every design choice to affected files, modules, storage, and validation paths",
            ],
            "prompt_types": ["architecture", "plan", "migration", "review"],
            "blueprints": ["architecture-spec", "repository-audit", "implementation-plan", "validation-audit"],
            "phases": ["discovery", "architecture", "planning"],
            "tags": ["role", "architecture", "design", "multi-agent"],
        },
        {
            "key": "role-senior-reviewer",
            "name": "Role: Senior Reviewer",
            "summary": "Frames the agent as a skeptical reviewer focused on findings, regressions, and missing proof.",
            "focus": [
                "identify the highest-risk flaws in {{artifact_or_plan_under_review}}",
                "prioritize bugs, weak assumptions, missing tests, and unsafe changes over style commentary",
                "force concrete evidence before accepting claims",
            ],
            "posture": [
                "present findings first and keep summary secondary",
                "cite the exact file, module, or behavior that creates the risk",
                "do not propose broad rewrites unless the current design is fundamentally unsafe",
            ],
            "prompt_types": ["review", "validation", "architecture", "security"],
            "blueprints": ["senior-code-review", "validation-audit", "security-hardening", "performance-hardening"],
            "phases": ["verification", "delivery"],
            "tags": ["role", "review", "findings-first", "risk"],
        },
        {
            "key": "role-implementation-planner",
            "name": "Role: Implementation Planner",
            "summary": "Frames the session as planning work that converts design into milestones, file maps, and checklists.",
            "focus": [
                "turn {{approved_architecture_or_goal}} into a step-by-step execution plan",
                "sequence dependencies so the implementer always has the next safe move",
                "make required tests, docs, migrations, and risk controls visible up front",
            ],
            "posture": [
                "plan in coherent slices that can be implemented and verified independently",
                "name likely files, modules, or repositories instead of speaking in abstractions only",
                "produce a plan that another agent could execute without rediscovering scope",
            ],
            "prompt_types": ["plan", "architecture", "migration", "implementation"],
            "blueprints": ["implementation-plan", "architecture-spec", "feature-implementation", "safe-refactor"],
            "phases": ["planning"],
            "tags": ["role", "planning", "checklists", "sequencing"],
        },
        {
            "key": "role-implementation-lead",
            "name": "Role: Implementation Lead",
            "summary": "Frames the session as hands-on delivery work focused on working code, tests, and incremental proof.",
            "focus": [
                "implement {{target_feature_or_fix}} end to end in the current repository",
                "keep changes small, coherent, and directly traceable to the stated goal",
                "produce runnable results rather than partial scaffolding",
            ],
            "posture": [
                "read existing code before editing it",
                "favor additive or low-risk refactors before deeper rewrites",
                "treat builds, tests, and manual verification as part of implementation rather than afterthoughts",
            ],
            "prompt_types": ["implementation", "bugfix", "refactor", "migration", "ui"],
            "blueprints": ["feature-implementation", "bugfix-with-regression-lock", "safe-refactor", "ui-ux-delivery", "embedded-firmware-iteration"],
            "phases": ["implementation", "verification"],
            "tags": ["role", "implementation", "delivery", "tests"],
        },
        {
            "key": "role-refactor-specialist",
            "name": "Role: Refactor Specialist",
            "summary": "Frames the session as safe refactoring with behavior preservation and regression locking.",
            "focus": [
                "improve the structure of {{target_area}} without changing externally expected behavior",
                "reduce complexity, coupling, or duplication while keeping the system stable",
                "use tests and checkpoints to prove the refactor did not drift into feature creep",
            ],
            "posture": [
                "start by reproducing current behavior and locking it with tests or fixtures",
                "prefer extracting seams and additive helpers before replacing large codepaths",
                "call out any behavior change explicitly instead of smuggling it in as refactoring",
            ],
            "prompt_types": ["refactor", "review", "testing", "implementation"],
            "blueprints": ["safe-refactor", "bugfix-with-regression-lock", "test-strategy-and-automation", "validation-audit"],
            "phases": ["discovery", "planning", "implementation", "verification"],
            "tags": ["role", "refactor", "regression", "stability"],
        },
        {
            "key": "role-test-validation-lead",
            "name": "Role: Test and Validation Lead",
            "summary": "Frames the session as a proof-gathering exercise covering automated and manual verification.",
            "focus": [
                "turn {{change_or_artifact}} into an explicit validation plan with evidence",
                "separate assumptions from verified facts",
                "find the cheapest set of tests that still gives high confidence",
            ],
            "posture": [
                "choose tests based on failure modes, not habit",
                "collect commands, fixtures, screenshots, logs, or traces as evidence",
                "be explicit about what could not be validated in the current environment",
            ],
            "prompt_types": ["testing", "validation", "review", "performance"],
            "blueprints": ["test-strategy-and-automation", "validation-audit", "performance-hardening", "feature-implementation"],
            "phases": ["verification", "delivery"],
            "tags": ["role", "validation", "testing", "evidence"],
        },
        {
            "key": "role-ui-ux-engineer",
            "name": "Role: UI and UX Engineer",
            "summary": "Frames the session around intentional interaction design, usability, and implementation realism.",
            "focus": [
                "design or refine {{ui_surface_or_flow}} so it is usable, credible, and implementation-ready",
                "connect information architecture, layout, states, and component responsibilities",
                "avoid placeholder UX that cannot survive real data and edge cases",
            ],
            "posture": [
                "preserve the product's visual language unless the prompt explicitly asks for redesign",
                "spell out interactions, empty states, error states, and responsiveness",
                "keep the design tied to the actual component system and stack",
            ],
            "prompt_types": ["ui", "architecture", "implementation", "review"],
            "blueprints": ["ui-ux-delivery", "architecture-spec", "validation-audit", "feature-implementation"],
            "phases": ["architecture", "planning", "implementation", "verification"],
            "tags": ["role", "ui", "ux", "interaction"],
        },
        {
            "key": "role-embedded-midi-engineer",
            "name": "Role: Embedded and MIDI Engineer",
            "summary": "Frames the session for hardware, firmware, MIDI, timing, and telemetry work.",
            "focus": [
                "improve {{firmware_or_realtime_pipeline}} with stable timing, observability, and safe hardware assumptions",
                "treat power, GPIO, memory, latency, and calibration as first-class constraints",
                "connect firmware changes to any host-side protocol, tooling, or UI surfaces that depend on them",
            ],
            "posture": [
                "avoid hand-wavy hardware guidance and tie recommendations to specific pins, buses, or timing paths",
                "prefer deterministic state machines and measurable thresholds over magic constants",
                "use fixtures, logs, or telemetry traces when real hardware is unavailable",
            ],
            "prompt_types": ["embedded", "implementation", "review", "testing"],
            "blueprints": ["embedded-firmware-iteration", "validation-audit", "feature-implementation", "test-strategy-and-automation"],
            "phases": ["discovery", "planning", "implementation", "verification"],
            "tags": ["role", "embedded", "midi", "firmware", "realtime"],
        },
    ]

    result: list[ComponentDefinition] = []
    for spec in role_specs:
        template = f"""
        ## Role
        You are acting as the {spec["name"].replace("Role: ", "").lower()} for this session.

        Primary responsibility:
        - {spec["focus"][0]}
        - {spec["focus"][1]}
        - {spec["focus"][2]}

        Working posture:
        - {spec["posture"][0]}
        - {spec["posture"][1]}
        - {spec["posture"][2]}
        """
        result.append(
            component(
                key=spec["key"],
                name=spec["name"],
                group="session-framing",
                block_kind="Instruction",
                summary=spec["summary"],
                template=template,
                tags=stable_tags(*spec["tags"]),
                prompt_types=spec["prompt_types"],
                blueprints=spec["blueprints"],
                phases=spec["phases"],
                recommended=True,
            )
        )
    return result


def build_mission_scope_components() -> list[ComponentDefinition]:
    shared_blueprints = [
        "architecture-spec",
        "repository-audit",
        "implementation-plan",
        "feature-implementation",
        "safe-refactor",
        "bugfix-with-regression-lock",
        "validation-audit",
        "ui-ux-delivery",
        "embedded-firmware-iteration",
    ]

    return [
        component(
            key="mission-exact-goal",
            name="Mission: Exact Goal",
            group="mission-scope",
            block_kind="Instruction",
            summary="Pins the session to one exact objective and prevents prompt drift.",
            template="""
            ## Mission
            Your exact goal is to {{exact_goal}}.

            Treat this as the primary objective for the session.
            Do not drift into adjacent improvements unless they are required to make {{exact_goal}} work correctly.
            If you discover a prerequisite, state it briefly and complete it before returning to the main objective.
            """,
            tags=stable_tags("mission", "goal", "scope"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=ALL_PHASES,
            recommended=True,
        ),
        component(
            key="mission-business-context",
            name="Mission: Business Context",
            group="mission-scope",
            block_kind="Instruction",
            summary="Adds the product or user reason behind the task so technical choices stay aligned.",
            template="""
            ## Why This Matters
            This work matters because {{business_context}}.

            Optimize for the user or business outcome, not for elegant but irrelevant technical changes.
            If two solutions are both correct, prefer the one that better supports {{business_context}}.
            """,
            tags=stable_tags("mission", "business-context", "prioritization"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=ALL_PHASES,
        ),
        component(
            key="scope-in-scope-items",
            name="Scope: In-Scope Items",
            group="mission-scope",
            block_kind="Constraint",
            summary="Defines the work that must be covered in the session.",
            template="""
            ## In Scope
            You are responsible for the following items:
            - {{in_scope_item_1}}
            - {{in_scope_item_2}}
            - {{in_scope_item_3}}

            Finish the in-scope slice end to end instead of leaving disconnected scaffolding.
            """,
            tags=stable_tags("scope", "in-scope", "requirements"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=ALL_PHASES,
            recommended=True,
        ),
        component(
            key="scope-out-of-scope-items",
            name="Scope: Out-of-Scope Items",
            group="mission-scope",
            block_kind="Constraint",
            summary="Prevents the model from wasting time on adjacent work or overbuilding.",
            template="""
            ## Out of Scope
            Do not spend time on the following unless they become mandatory blockers:
            - {{out_of_scope_item_1}}
            - {{out_of_scope_item_2}}
            - {{out_of_scope_item_3}}

            If you encounter one of these areas, acknowledge it and return to the main objective.
            """,
            tags=stable_tags("scope", "out-of-scope", "focus"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=ALL_PHASES,
            recommended=True,
        ),
        component(
            key="success-criteria",
            name="Success Criteria",
            group="mission-scope",
            block_kind="Validation",
            summary="Defines what must be true before the work is considered complete.",
            template="""
            ## Success Criteria
            The work is complete only when all of the following are true:
            - {{success_criterion_1}}
            - {{success_criterion_2}}
            - {{success_criterion_3}}

            Do not claim completion if any criterion is still assumed rather than verified.
            """,
            tags=stable_tags("success", "acceptance", "definition-of-done"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=["planning", "implementation", "verification", "delivery"],
            recommended=True,
        ),
        component(
            key="required-deliverables",
            name="Required Deliverables",
            group="mission-scope",
            block_kind="Delivery",
            summary="Forces the prompt to name the concrete outputs that must be produced.",
            template="""
            ## Required Deliverables
            Produce the following deliverables in this session:
            - {{deliverable_1}}
            - {{deliverable_2}}
            - {{deliverable_3}}

            The session is not complete if the implementation is finished but the required artifacts are missing.
            """,
            tags=stable_tags("deliverables", "artifacts", "outputs"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=["planning", "implementation", "verification", "delivery"],
            recommended=True,
        ),
        component(
            key="first-response-contract",
            name="First Response Contract",
            group="mission-scope",
            block_kind="Delivery",
            summary="Controls what the agent must include in its very first response before deeper work starts.",
            template="""
            ## First Response Contract
            Your first response must contain:
            - a short restatement of the task,
            - the first files, artifacts, or systems you will inspect,
            - the immediate risks or assumptions that could change the approach.

            Keep that first response short and operational. Do not jump into broad design prose before confirming the workspace.
            """,
            tags=stable_tags("first-response", "workflow", "contract"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=["discovery"],
        ),
        component(
            key="stop-condition",
            name="Stop Condition",
            group="mission-scope",
            block_kind="Constraint",
            summary="Defines when the agent must stop instead of continuing automatically.",
            template="""
            ## Stop Condition
            Stop when {{stop_condition}}.

            Do not silently continue into the next milestone once this point is reached.
            When you stop, summarize what is complete, what remains, and the recommended next prompt or next agent.
            """,
            tags=stable_tags("stop-condition", "phase-gate", "handoff"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=shared_blueprints,
            phases=["architecture", "planning", "implementation", "verification", "delivery"],
        ),
    ]


def build_context_discovery_components() -> list[ComponentDefinition]:
    broad_blueprints = [
        "architecture-spec",
        "repository-audit",
        "implementation-plan",
        "feature-implementation",
        "safe-refactor",
        "bugfix-with-regression-lock",
        "senior-code-review",
        "validation-audit",
        "ui-ux-delivery",
        "embedded-firmware-iteration",
    ]
    return [
        component(
            key="repo-map-confirmation",
            name="Context: Repository Map Confirmation",
            group="context-discovery",
            block_kind="Instruction",
            summary="Requires the agent to confirm the real project structure instead of assuming paths from the prompt.",
            template="""
            ## Repository Map
            Confirm the repository structure before making changes.

            At minimum, verify:
            - {{solution_or_workspace_root}}
            - {{primary_projects_or_modules}}
            - {{tests_and_validation_projects}}
            - {{docs_or_artifact_paths}}

            If the working tree differs from the prompt, resolve the mismatch explicitly instead of ignoring it.
            """,
            tags=stable_tags("repository-map", "paths", "context"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=["discovery"],
            recommended=True,
        ),
        component(
            key="required-reading-list",
            name="Context: Required Reading List",
            group="context-discovery",
            block_kind="Instruction",
            summary="Lists the files, docs, and artifacts the agent must read before acting.",
            template="""
            ## Inputs to Read
            Read the following before you plan or implement:
            - {{input_path_1}}
            - {{input_path_2}}
            - {{input_path_3}}
            - {{input_path_4}}

            Treat these inputs as authoritative unless the codebase proves they are outdated.
            """,
            tags=stable_tags("inputs", "required-reading", "artifacts"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=["discovery"],
            recommended=True,
        ),
        component(
            key="current-state-audit",
            name="Context: Current State Audit",
            group="context-discovery",
            block_kind="Instruction",
            summary="Makes the agent audit what already exists before proposing changes or declaring gaps.",
            template="""
            ## Current State Audit
            Start by auditing what already exists for {{target_area}}.

            Confirm:
            - what is already implemented,
            - what is partial or inconsistent,
            - what is missing,
            - what existing tests or fixtures already cover.

            Do not propose new architecture or new files before you know the real baseline.
            """,
            tags=stable_tags("audit", "baseline", "discovery"),
            prompt_types=["architecture", "audit", "plan", "implementation", "refactor", "bugfix", "review", "validation"],
            blueprints=broad_blueprints,
            phases=["discovery"],
            recommended=True,
        ),
        component(
            key="file-touch-plan",
            name="Context: File Touch Plan",
            group="context-discovery",
            block_kind="Instruction",
            summary="Requires the agent to predict the likely files and modules involved before editing begins.",
            template="""
            ## File Touch Plan
            Before editing, identify the files or modules most likely to change:
            - {{likely_file_or_module_1}}
            - {{likely_file_or_module_2}}
            - {{likely_file_or_module_3}}

            Call out any high-risk files where careful review is required because regressions would be expensive.
            """,
            tags=stable_tags("file-plan", "impact-analysis", "change-scope"),
            prompt_types=["architecture", "plan", "implementation", "refactor", "bugfix", "review"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "senior-code-review"],
            phases=["discovery", "planning"],
        ),
        component(
            key="dependency-inventory",
            name="Context: Dependency Inventory",
            group="context-discovery",
            block_kind="Instruction",
            summary="Makes the agent identify runtime, framework, and integration dependencies that constrain the work.",
            template="""
            ## Dependency Inventory
            List the important dependencies that shape this work:
            - frameworks and runtime versions,
            - external services or protocols,
            - build, test, and deployment dependencies,
            - hardware or browser capabilities if relevant.

            If a dependency version or capability is uncertain, verify it before designing around it.
            """,
            tags=stable_tags("dependencies", "stack", "integration"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=["discovery", "planning"],
        ),
        component(
            key="environment-and-commands",
            name="Context: Environment and Commands",
            group="context-discovery",
            block_kind="Instruction",
            summary="Pins the build, run, and test commands that should be used during the session.",
            template="""
            ## Local Build and Verification Commands
            Use and update the real commands for this workspace:
            - build: {{build_command}}
            - unit tests: {{unit_test_command}}
            - integration tests: {{integration_test_command}}
            - UI tests: {{ui_test_command}}
            - run app or service: {{run_command}}

            If these commands are wrong for the current workspace, correct them early and keep the corrected commands visible in the session.
            """,
            tags=stable_tags("commands", "environment", "verification"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=["discovery", "implementation", "verification"],
            recommended=True,
        ),
        component(
            key="relevant-artifacts-and-fixtures",
            name="Context: Relevant Artifacts and Fixtures",
            group="context-discovery",
            block_kind="Instruction",
            summary="Makes the agent locate fixtures, examples, sample data, screenshots, or diagrams before building from scratch.",
            template="""
            ## Existing Artifacts and Fixtures
            Look for existing artifacts that should guide or validate the work:
            - fixtures or sample data,
            - screenshots or recordings,
            - diagrams or architecture notes,
            - protocol examples or golden outputs.

            Reuse and extend those artifacts where possible instead of inventing disconnected new ones.
            """,
            tags=stable_tags("fixtures", "artifacts", "reuse"),
            prompt_types=["architecture", "plan", "implementation", "refactor", "bugfix", "testing", "ui", "embedded"],
            blueprints=broad_blueprints,
            phases=["discovery", "planning", "verification"],
        ),
        component(
            key="assumptions-and-open-questions",
            name="Context: Assumptions and Open Questions",
            group="context-discovery",
            block_kind="Delivery",
            summary="Captures the assumptions the agent is making and the questions that still matter.",
            template="""
            ## Assumptions and Open Questions
            Keep a short list of:
            - assumptions you are making because the repo or prompt is incomplete,
            - questions that would materially change the approach,
            - questions that can safely be deferred without blocking progress.

            Separate confirmed facts from inferred assumptions.
            """,
            tags=stable_tags("assumptions", "questions", "risk"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=ALL_PHASES,
        ),
    ]


def build_guardrail_components() -> list[ComponentDefinition]:
    broad_blueprints = [
        "architecture-spec",
        "repository-audit",
        "implementation-plan",
        "feature-implementation",
        "safe-refactor",
        "bugfix-with-regression-lock",
        "senior-code-review",
        "test-strategy-and-automation",
        "validation-audit",
        "performance-hardening",
        "security-hardening",
        "ui-ux-delivery",
        "embedded-firmware-iteration",
    ]
    return [
        component(
            key="non-negotiable-rules",
            name="Guardrail: Non-Negotiable Rules",
            group="guardrails",
            block_kind="Constraint",
            summary="Provides hard rules that the agent must not violate.",
            template="""
            ## Non-Negotiable Rules
            The following rules are mandatory:
            - {{rule_1}}
            - {{rule_2}}
            - {{rule_3}}
            - {{rule_4}}

            If a rule conflicts with a proposed solution, change the solution instead of relaxing the rule silently.
            """,
            tags=stable_tags("guardrails", "hard-rules", "constraints"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=ALL_PHASES,
            recommended=True,
        ),
        component(
            key="preserve-architecture-boundaries",
            name="Guardrail: Preserve Architecture Boundaries",
            group="guardrails",
            block_kind="Constraint",
            summary="Prevents the agent from smearing business logic across layers or breaking modular boundaries.",
            template="""
            ## Architecture Boundaries
            Preserve the existing architectural boundaries unless the prompt explicitly authorizes a redesign.

            Do not:
            - move business logic into UI-only code,
            - bypass shared contracts or data access patterns,
            - introduce cross-module coupling just to finish faster.

            If a boundary is wrong, document the issue and fix it deliberately rather than by accident.
            """,
            tags=stable_tags("architecture", "boundaries", "maintainability"),
            prompt_types=["architecture", "implementation", "refactor", "review", "security", "migration"],
            blueprints=broad_blueprints,
            phases=["architecture", "planning", "implementation", "verification"],
            recommended=True,
        ),
        component(
            key="small-verifiable-increments",
            name="Guardrail: Small Verifiable Increments",
            group="guardrails",
            block_kind="Constraint",
            summary="Forces stepwise changes with proof after each slice.",
            template="""
            ## Increment Size
            Work in small, verifiable increments.

            For each slice:
            - change only the minimum coherent surface,
            - run the closest relevant verification,
            - keep the system buildable and testable before moving on.

            If the work starts turning into a large refactor, split it into smaller steps with explicit checkpoints.
            """,
            tags=stable_tags("increments", "verification", "risk-reduction"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=["planning", "implementation", "verification"],
            recommended=True,
        ),
        component(
            key="comments-in-english",
            name="Guardrail: Comments in English",
            group="guardrails",
            block_kind="Constraint",
            summary="Keeps code comments and inline documentation in English for consistency.",
            template="""
            ## Comment Language
            All new or updated code comments must be in English.

            Prefer concise, useful comments that explain non-obvious behavior.
            Do not add commentary that only restates what the code already makes obvious.
            """,
            tags=stable_tags("comments", "english", "consistency"),
            prompt_types=["implementation", "refactor", "bugfix", "embedded", "ui"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["implementation"],
        ),
        component(
            key="no-placeholder-ui",
            name="Guardrail: No Placeholder-Only UI",
            group="guardrails",
            block_kind="Constraint",
            summary="Prevents shallow UI work that looks complete but is not wired to real state or behavior.",
            template="""
            ## No Placeholder-Only UI
            Do not leave placeholder-only UI in the finished result.

            New screens or controls must:
            - bind to real state and real workflows,
            - handle empty, loading, and error states where relevant,
            - use the actual component system and styling approach in the repo.
            """,
            tags=stable_tags("ui", "quality-bar", "real-behavior"),
            prompt_types=["ui", "implementation", "review"],
            blueprints=["ui-ux-delivery", "feature-implementation", "validation-audit", "senior-code-review"],
            phases=["implementation", "verification"],
        ),
        component(
            key="preserve-backward-compatibility",
            name="Guardrail: Preserve Backward Compatibility",
            group="guardrails",
            block_kind="Constraint",
            summary="Protects existing data, contracts, and user flows unless change is explicit and managed.",
            template="""
            ## Backward Compatibility
            Preserve existing contracts, persisted data, and user-visible behavior unless the prompt explicitly calls for a breaking change.

            If a breaking change is required:
            - call it out explicitly,
            - add migration or compatibility handling,
            - document the risk and the rollback path.
            """,
            tags=stable_tags("compatibility", "migrations", "contracts"),
            prompt_types=["architecture", "implementation", "refactor", "bugfix", "migration"],
            blueprints=["architecture-spec", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration"],
            phases=["planning", "implementation", "verification"],
            recommended=True,
        ),
        component(
            key="safe-ambiguity-handling",
            name="Guardrail: Safe Ambiguity Handling",
            group="guardrails",
            block_kind="Constraint",
            summary="Tells the agent how to behave when requirements or repo facts are ambiguous.",
            template="""
            ## Ambiguity Handling
            When the prompt or repository is ambiguous:
            - prefer the simplest behavior that stays consistent with the existing system,
            - document the assumption you chose,
            - add tests or notes that lock the decision in.

            Do not invent hidden requirements or silently make risky assumptions.
            """,
            tags=stable_tags("ambiguity", "assumptions", "safety"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=ALL_PHASES,
            recommended=True,
        ),
        component(
            key="protect-secrets-and-sensitive-data",
            name="Guardrail: Protect Secrets and Sensitive Data",
            group="guardrails",
            block_kind="Security",
            summary="Prevents secret leakage and unsafe handling of sensitive values.",
            template="""
            ## Secret and Sensitive Data Handling
            Do not expose secrets, tokens, private keys, or raw sensitive values in prompts, logs, screenshots, or generated code.

            If the work touches authentication, billing, telemetry, or user data:
            - keep secrets server-side when possible,
            - redact sensitive values from examples and output,
            - call out any unsafe storage or transport pattern you encounter.
            """,
            tags=stable_tags("security", "secrets", "privacy"),
            prompt_types=["architecture", "implementation", "review", "security", "validation"],
            blueprints=["architecture-spec", "feature-implementation", "senior-code-review", "security-hardening", "validation-audit"],
            phases=["architecture", "implementation", "verification", "delivery"],
            recommended=True,
        ),
        component(
            key="honest-blocker-reporting",
            name="Guardrail: Honest Blocker Reporting",
            group="guardrails",
            block_kind="Delivery",
            summary="Forces the agent to report environment, access, or test blockers honestly.",
            template="""
            ## Blocker Reporting
            If you are blocked by the environment, missing hardware, unavailable credentials, or failing infrastructure:
            - say exactly what is blocked,
            - describe what you verified before concluding it is blocked,
            - give the closest fallback or next action.

            Do not imply that blocked verification passed.
            """,
            tags=stable_tags("blockers", "honesty", "verification"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=broad_blueprints,
            phases=["verification", "delivery"],
            recommended=True,
        ),
        component(
            key="no-destructive-changes",
            name="Guardrail: No Destructive Changes",
            group="guardrails",
            block_kind="Constraint",
            summary="Prevents destructive commands or silent data loss while the agent works.",
            template="""
            ## Destructive Change Ban
            Do not use destructive commands or irreversible cleanup unless the prompt explicitly authorizes them.

            Avoid:
            - deleting user work or unreviewed changes,
            - force-resetting branches or databases,
            - throwing away fixtures, logs, or screenshots that still matter for proof.
            """,
            tags=stable_tags("safety", "git", "destructive-actions"),
            prompt_types=["implementation", "refactor", "bugfix", "review", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "senior-code-review", "embedded-firmware-iteration"],
            phases=["implementation", "verification"],
            recommended=True,
        ),
    ]


def build_workflow_components() -> list[ComponentDefinition]:
    return [
        component(
            key="sequential-phases",
            name="Workflow: Sequential Phases",
            group="workflow-orchestration",
            block_kind="Instruction",
            summary="Defines an explicit phase order so the agent does not merge discovery, planning, implementation, and validation into one blur.",
            template="""
            ## Phase Order
            Execute this work in order:
            1. discovery and audit
            2. architecture or plan
            3. implementation
            4. verification and review
            5. delivery or handoff

            Do not collapse all phases into a single response if phase-specific proof is required.
            """,
            tags=stable_tags("workflow", "phases", "sequencing"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "repository-audit", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "embedded-firmware-iteration"],
            phases=ALL_PHASES,
            recommended=True,
        ),
        component(
            key="stop-after-phase",
            name="Workflow: Stop After Phase",
            group="workflow-orchestration",
            block_kind="Constraint",
            summary="Turns each phase into a gate and prevents the agent from rolling into the next one automatically.",
            template="""
            ## Phase Gate
            After completing the current phase:
            - run the required validation for that phase,
            - summarize what changed,
            - stop and wait for the next prompt or next instruction if this workflow is phase-gated.

            Do not continue automatically into the next phase when the workflow is meant to be reviewed between steps.
            """,
            tags=stable_tags("workflow", "gate", "stop-rule"),
            prompt_types=["architecture", "plan", "implementation", "refactor", "embedded", "ui", "validation"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["architecture", "planning", "implementation", "verification", "delivery"],
            recommended=True,
        ),
        component(
            key="multi-agent-chain",
            name="Workflow: Multi-Agent Chain",
            group="workflow-orchestration",
            block_kind="Instruction",
            summary="Defines the canonical chain of architecture, review, planning, implementation, and validation agents.",
            template="""
            ## Multi-Agent Chain
            Organize this workflow as a chain of specialized agents:
            1. architecture agent creates the design and constraints,
            2. reviewer challenges the design and exposes risks,
            3. planning agent converts it into milestones and checklists,
            4. implementation agent delivers the code in slices,
            5. validation agent gathers proof and performs the final audit.

            Each downstream agent must inherit the outputs and unresolved risks from the previous one instead of rediscovering context from scratch.
            """,
            tags=stable_tags("multi-agent", "handoff", "workflow"),
            prompt_types=["architecture", "plan", "implementation", "review", "validation", "migration"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "validation-audit"],
            phases=ALL_PHASES,
            recommended=True,
        ),
        component(
            key="checklist-update-loop",
            name="Workflow: Checklist Update Loop",
            group="workflow-orchestration",
            block_kind="Instruction",
            summary="Requires the agent to maintain the checklist as work progresses.",
            template="""
            ## Checklist Update Loop
            Maintain the active checklist during the session.

            After each meaningful step:
            - mark completed items,
            - note what remains,
            - record the validation that justified the state change.

            Never report an item as done without naming the evidence.
            """,
            tags=stable_tags("checklist", "progress", "evidence"),
            prompt_types=["plan", "implementation", "refactor", "bugfix", "testing", "validation", "embedded"],
            blueprints=["implementation-plan", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "test-strategy-and-automation", "validation-audit", "embedded-firmware-iteration"],
            phases=["planning", "implementation", "verification", "delivery"],
        ),
        component(
            key="persistent-progress-files",
            name="Workflow: Persistent Progress Files",
            group="workflow-orchestration",
            block_kind="Delivery",
            summary="Creates durable progress artifacts so future sessions can continue without losing context.",
            template="""
            ## Persistent Progress
            Keep persistent progress artifacts in-repo so later sessions can resume cleanly.

            Maintain:
            - a status file,
            - a short decisions log,
            - a known-gaps or postponed-items log,
            - a next-prompt or next-step pointer.
            """,
            tags=stable_tags("progress", "continuity", "handoff"),
            prompt_types=["implementation", "refactor", "bugfix", "migration", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration", "validation-audit"],
            phases=["planning", "implementation", "verification", "delivery"],
        ),
        component(
            key="decision-log-maintenance",
            name="Workflow: Decision Log Maintenance",
            group="workflow-orchestration",
            block_kind="Delivery",
            summary="Tracks decisions that future agents should not have to rediscover.",
            template="""
            ## Decision Log
            Record decisions that affect future work:
            - what decision was made,
            - why it was chosen,
            - what tradeoff or limitation remains,
            - what evidence supports the choice.

            Keep the log short and operational, not essay-like.
            """,
            tags=stable_tags("decisions", "adr", "continuity"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "validation-audit", "embedded-firmware-iteration"],
            phases=["architecture", "planning", "implementation", "verification", "delivery"],
        ),
        component(
            key="known-gaps-log",
            name="Workflow: Known Gaps Log",
            group="workflow-orchestration",
            block_kind="Delivery",
            summary="Separates intentionally deferred work from unfinished or forgotten work.",
            template="""
            ## Known Gaps
            Keep a short known-gaps list for items that are deliberately postponed.

            For each gap, note:
            - why it was not completed now,
            - what risk it creates,
            - what event should bring it back into scope.
            """,
            tags=stable_tags("known-gaps", "deferred-work", "risk"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "embedded-firmware-iteration"],
            phases=["planning", "implementation", "verification", "delivery"],
        ),
        component(
            key="next-prompt-pointer",
            name="Workflow: Next Prompt Pointer",
            group="workflow-orchestration",
            block_kind="Delivery",
            summary="Requires the agent to say which next prompt or next flow step should run.",
            template="""
            ## Next Prompt Pointer
            End the current phase by naming the recommended next prompt, next phase, or next agent.

            The pointer should explain:
            - what is now ready,
            - what the next step should consume,
            - what unresolved risks the next step must carry forward.
            """,
            tags=stable_tags("next-step", "handoff", "workflow"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["architecture", "planning", "implementation", "verification", "delivery"],
        ),
        component(
            key="branch-and-status-tracking",
            name="Workflow: Branch and Status Tracking",
            group="workflow-orchestration",
            block_kind="Instruction",
            summary="Makes branch identity, run status, and working tree state visible in the workflow.",
            template="""
            ## Branch and Status Tracking
            Keep branch and session status explicit:
            - current branch or branch plan,
            - current milestone or phase,
            - whether the working tree is clean or contains relevant in-flight changes.

            If the work depends on dirty state, describe how you will avoid clobbering unrelated changes.
            """,
            tags=stable_tags("branch", "status", "git", "workflow"),
            prompt_types=["implementation", "refactor", "bugfix", "review", "migration", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "senior-code-review", "embedded-firmware-iteration"],
            phases=["discovery", "planning", "implementation", "verification"],
        ),
        component(
            key="required-phase-output-format",
            name="Workflow: Required Phase Output Format",
            group="workflow-orchestration",
            block_kind="Delivery",
            summary="Standardizes what each phase summary must contain.",
            template="""
            ## Phase Output Format
            At the end of the phase, provide:
            1. what you inspected,
            2. what you changed or produced,
            3. what validation you ran,
            4. remaining risks or blockers,
            5. the recommended next step.

            Keep the format consistent so later agents and reviewers can diff progress quickly.
            """,
            tags=stable_tags("output-format", "phase-summary", "handoff"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["architecture", "planning", "implementation", "verification", "delivery"],
            recommended=True,
        ),
    ]


def build_architecture_analysis_components() -> list[ComponentDefinition]:
    return [
        component(
            key="architecture-blueprint",
            name="Architecture: Blueprint",
            group="architecture-analysis",
            block_kind="Instruction",
            summary="Requests an implementation-ready architecture with modules, flows, risks, and proof paths.",
            template="""
            ## Architecture Blueprint
            Produce an implementation-ready architecture for {{target_feature_or_system}}.

            Cover:
            - module or service boundaries,
            - data flow and control flow,
            - storage or state ownership,
            - external interfaces or contracts,
            - validation strategy and major risks.

            The blueprint must be specific enough that another agent could implement it without redesigning it.
            """,
            tags=stable_tags("architecture", "blueprint", "design"),
            prompt_types=["architecture", "plan", "migration", "review"],
            blueprints=["architecture-spec", "implementation-plan", "validation-audit"],
            phases=["architecture"],
            recommended=True,
        ),
        component(
            key="gap-analysis",
            name="Architecture: Gap Analysis",
            group="architecture-analysis",
            block_kind="Instruction",
            summary="Compares the desired target with the current implementation and names the real gaps.",
            template="""
            ## Gap Analysis
            Compare the current implementation with the target outcome for {{target_feature_or_system}}.

            Identify:
            - what is already done,
            - what is partial or inconsistent,
            - what is missing entirely,
            - what creates the largest delivery or quality risk.

            Focus on actionable gaps rather than vague observations.
            """,
            tags=stable_tags("gap-analysis", "audit", "planning"),
            prompt_types=["architecture", "audit", "plan", "review", "migration"],
            blueprints=["architecture-spec", "repository-audit", "implementation-plan", "validation-audit"],
            phases=["discovery", "architecture", "planning"],
            recommended=True,
        ),
        component(
            key="parity-matrix",
            name="Architecture: Parity Matrix",
            group="architecture-analysis",
            block_kind="Instruction",
            summary="Creates an explicit mapping between current assets and target assets for migrations or rebuilds.",
            template="""
            ## Parity Matrix
            Build a parity matrix for {{source_system}} to {{target_system}}.

            For each in-scope page, route, module, or workflow, map:
            - the current implementation owner,
            - the target implementation owner,
            - reusable assets or contracts,
            - missing work,
            - required tests.
            """,
            tags=stable_tags("parity", "migration", "mapping"),
            prompt_types=["architecture", "audit", "plan", "migration"],
            blueprints=["architecture-spec", "repository-audit", "implementation-plan"],
            phases=["discovery", "architecture", "planning"],
        ),
        component(
            key="domain-model-and-entities",
            name="Architecture: Domain Model and Entities",
            group="architecture-analysis",
            block_kind="Instruction",
            summary="Requires explicit domain concepts, entities, aggregates, or records before implementation starts.",
            template="""
            ## Domain Model
            Define the domain model needed for {{target_feature_or_problem}}.

            Make explicit:
            - key entities or records,
            - important identifiers and relationships,
            - lifecycle states,
            - invariants that must remain true.

            Keep the model aligned with the existing system language instead of inventing a second vocabulary.
            """,
            tags=stable_tags("domain-model", "entities", "data"),
            prompt_types=["architecture", "plan", "implementation", "migration"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "embedded-firmware-iteration"],
            phases=["architecture", "planning"],
        ),
        component(
            key="api-contract-design",
            name="Architecture: API Contract Design",
            group="architecture-analysis",
            block_kind="Instruction",
            summary="Defines endpoints, messages, DTOs, protocol changes, or interop contracts before coding.",
            template="""
            ## API or Protocol Contract
            Define the contract for {{api_or_protocol_surface}}.

            Include:
            - operations or messages,
            - payload shape,
            - validation rules,
            - versioning or compatibility notes,
            - error handling and observability expectations.
            """,
            tags=stable_tags("api", "contract", "protocol", "interop"),
            prompt_types=["architecture", "plan", "implementation", "migration", "embedded"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "embedded-firmware-iteration"],
            phases=["architecture", "planning"],
        ),
        component(
            key="data-model-and-migration-design",
            name="Architecture: Data Model and Migration Design",
            group="architecture-analysis",
            block_kind="Instruction",
            summary="Covers schema, persistence, migrations, and compatibility risks before data changes are made.",
            template="""
            ## Data Model and Migration Design
            Design the persistence changes for {{data_change_or_feature}}.

            Cover:
            - tables, documents, or persisted records affected,
            - indexes and query paths,
            - migration or seed strategy,
            - backward compatibility and rollback concerns,
            - test coverage needed across supported databases.
            """,
            tags=stable_tags("data-model", "migration", "schema", "persistence"),
            prompt_types=["architecture", "plan", "implementation", "migration", "performance"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "performance-hardening"],
            phases=["architecture", "planning"],
        ),
        component(
            key="ux-surface-map",
            name="Architecture: UX Surface Map",
            group="architecture-analysis",
            block_kind="Instruction",
            summary="Maps the screens, states, interactions, and component responsibilities for UI work.",
            template="""
            ## UX Surface Map
            Map the user-facing surfaces for {{target_ui_flow}}.

            Show:
            - main pages, tabs, or panels,
            - primary actions,
            - state transitions,
            - error, loading, and empty states,
            - where reusable components should own shared behavior.
            """,
            tags=stable_tags("ux", "screens", "component-map"),
            prompt_types=["architecture", "ui", "plan", "review"],
            blueprints=["ui-ux-delivery", "architecture-spec", "implementation-plan", "validation-audit"],
            phases=["architecture", "planning"],
        ),
        component(
            key="risk-register",
            name="Architecture: Risk Register",
            group="architecture-analysis",
            block_kind="Validation",
            summary="Makes the agent name the major risks, failure modes, and mitigations before implementation.",
            template="""
            ## Risk Register
            List the main risks for {{target_feature_or_change}}.

            For each risk, capture:
            - the failure mode,
            - why it matters,
            - the mitigation,
            - the validation that will prove the mitigation worked.
            """,
            tags=stable_tags("risk", "mitigation", "planning"),
            prompt_types=["architecture", "plan", "review", "validation", "security", "performance", "embedded"],
            blueprints=["architecture-spec", "implementation-plan", "validation-audit", "security-hardening", "performance-hardening", "embedded-firmware-iteration"],
            phases=["architecture", "planning", "verification"],
            recommended=True,
        ),
    ]


def build_planning_components() -> list[ComponentDefinition]:
    return [
        component(
            key="implementation-plan-step-by-step",
            name="Planning: Step-by-Step Implementation Plan",
            group="planning-checklists",
            block_kind="Instruction",
            summary="Turns the target work into a concrete sequence of steps with dependencies and proof.",
            template="""
            ## Implementation Plan
            Create a step-by-step implementation plan for {{target_feature_or_fix}}.

            For each step, include:
            - the objective,
            - files or modules likely involved,
            - the validation that will prove the step is complete,
            - any dependency on an earlier step.
            """,
            tags=stable_tags("implementation-plan", "sequence", "checklist"),
            prompt_types=["plan", "architecture", "implementation", "migration"],
            blueprints=["implementation-plan", "architecture-spec", "feature-implementation", "safe-refactor", "embedded-firmware-iteration"],
            phases=["planning"],
            recommended=True,
        ),
        component(
            key="milestone-breakdown",
            name="Planning: Milestone Breakdown",
            group="planning-checklists",
            block_kind="Instruction",
            summary="Groups steps into meaningful milestones so work can be reviewed between major slices.",
            template="""
            ## Milestones
            Break the work into milestones that are meaningful review points.

            Each milestone should:
            - have a clear goal,
            - remain small enough to validate,
            - leave the repository in a stable state,
            - make the next milestone easier instead of more ambiguous.
            """,
            tags=stable_tags("milestones", "planning", "phases"),
            prompt_types=["plan", "architecture", "implementation", "migration"],
            blueprints=["implementation-plan", "architecture-spec", "feature-implementation", "embedded-firmware-iteration"],
            phases=["planning"],
        ),
        component(
            key="dependency-ordering",
            name="Planning: Dependency Ordering",
            group="planning-checklists",
            block_kind="Instruction",
            summary="Forces the plan to respect technical dependencies instead of random task ordering.",
            template="""
            ## Dependency Ordering
            Order the work by dependency depth.

            Start with:
            - shared contracts and foundational abstractions,
            - storage or protocol changes,
            - thin wiring layers,
            - user-facing surfaces and tests after the foundation exists.

            If the order must differ, explain why.
            """,
            tags=stable_tags("dependencies", "ordering", "planning"),
            prompt_types=["plan", "architecture", "implementation", "migration"],
            blueprints=["implementation-plan", "architecture-spec", "feature-implementation", "safe-refactor"],
            phases=["planning"],
            recommended=True,
        ),
        component(
            key="files-and-modules-likely-involved",
            name="Planning: Files and Modules Likely Involved",
            group="planning-checklists",
            block_kind="Instruction",
            summary="Requires file-level or module-level targeting in the plan.",
            template="""
            ## Files and Modules Likely Involved
            For each planned step, name the files, directories, modules, or services most likely to be touched.

            Avoid vague plans such as "update the backend" or "fix the UI".
            The goal is to reduce rediscovery work for the implementer.
            """,
            tags=stable_tags("files", "modules", "planning"),
            prompt_types=["plan", "implementation", "refactor", "bugfix", "migration"],
            blueprints=["implementation-plan", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration"],
            phases=["planning"],
        ),
        component(
            key="test-plan-matrix",
            name="Planning: Test Plan Matrix",
            group="planning-checklists",
            block_kind="Testing",
            summary="Maps each planned change to the tests or validation evidence it needs.",
            template="""
            ## Test Plan Matrix
            Build a test matrix for the planned work.

            For each major step, define:
            - unit tests,
            - integration or contract tests,
            - UI or end-to-end tests,
            - manual checks if automation is not enough.

            Tie each test back to a specific failure mode.
            """,
            tags=stable_tags("test-plan", "matrix", "quality"),
            prompt_types=["plan", "testing", "validation", "implementation"],
            blueprints=["implementation-plan", "test-strategy-and-automation", "feature-implementation", "validation-audit"],
            phases=["planning", "verification"],
            recommended=True,
        ),
        component(
            key="definition-of-done",
            name="Planning: Definition of Done",
            group="planning-checklists",
            block_kind="Validation",
            summary="Creates an explicit quality bar for the work.",
            template="""
            ## Definition of Done
            Define what "done" means for {{target_feature_or_change}}.

            Cover:
            - functional correctness,
            - testing and evidence,
            - documentation or artifact updates,
            - performance, accessibility, or security expectations where relevant,
            - any environment-specific proof that must exist.
            """,
            tags=stable_tags("definition-of-done", "quality-bar", "acceptance"),
            prompt_types=["plan", "architecture", "validation", "implementation"],
            blueprints=["implementation-plan", "architecture-spec", "validation-audit", "test-strategy-and-automation"],
            phases=["planning", "verification", "delivery"],
            recommended=True,
        ),
        component(
            key="acceptance-checklist",
            name="Planning: Acceptance Checklist",
            group="planning-checklists",
            block_kind="Validation",
            summary="Turns success criteria into checklist form that can be updated during execution.",
            template="""
            ## Acceptance Checklist
            Convert the success criteria into a checklist that can be updated during delivery.

            Each checklist item should be:
            - binary enough to confirm,
            - tied to a validation step,
            - specific enough that a reviewer can challenge it.
            """,
            tags=stable_tags("acceptance", "checklist", "verification"),
            prompt_types=["plan", "implementation", "testing", "validation"],
            blueprints=["implementation-plan", "feature-implementation", "test-strategy-and-automation", "validation-audit"],
            phases=["planning", "implementation", "verification", "delivery"],
        ),
        component(
            key="rollback-and-recovery-plan",
            name="Planning: Rollback and Recovery Plan",
            group="planning-checklists",
            block_kind="Validation",
            summary="Requires the plan to include how to back out or recover from a bad change.",
            template="""
            ## Rollback and Recovery
            Define the rollback or recovery path if {{target_feature_or_change}} fails in validation or rollout.

            Include:
            - what can be reverted safely,
            - what data or schema risk exists,
            - what fallback behavior should remain available,
            - what evidence would trigger rollback.
            """,
            tags=stable_tags("rollback", "recovery", "risk"),
            prompt_types=["plan", "implementation", "migration", "performance", "security"],
            blueprints=["implementation-plan", "feature-implementation", "security-hardening", "performance-hardening", "embedded-firmware-iteration"],
            phases=["planning", "verification", "delivery"],
        ),
    ]


def build_implementation_components() -> list[ComponentDefinition]:
    return [
        component(
            key="implement-in-small-slices",
            name="Implementation: Small Safe Slices",
            group="implementation-execution",
            block_kind="Instruction",
            summary="Requires the implementation to proceed in small slices with proof after each one.",
            template="""
            ## Implementation Style
            Implement the work in small, safe slices.

            Each slice should:
            - change one coherent behavior or structural step,
            - keep the codebase buildable,
            - be followed immediately by the closest relevant validation.
            """,
            tags=stable_tags("implementation", "slices", "verification"),
            prompt_types=["implementation", "refactor", "bugfix", "migration", "embedded", "ui"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["implementation", "verification"],
            recommended=True,
        ),
        component(
            key="additive-refactor-first",
            name="Implementation: Additive Refactor First",
            group="implementation-execution",
            block_kind="Instruction",
            summary="Encourages extraction and additive seams before replacing core behavior.",
            template="""
            ## Additive Refactor First
            When the work needs structural change, prefer this order:
            1. introduce the new helper, contract, or seam,
            2. wire it into the existing code with minimal behavior change,
            3. switch behavior only after tests cover the new path.

            Avoid large one-step rewrites unless the target area is already isolated and well tested.
            """,
            tags=stable_tags("refactor", "additive", "risk-reduction"),
            prompt_types=["implementation", "refactor", "bugfix", "migration"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "performance-hardening"],
            phases=["implementation", "verification"],
            recommended=True,
        ),
        component(
            key="keep-public-surface-minimal",
            name="Implementation: Keep Public Surface Minimal",
            group="implementation-execution",
            block_kind="Constraint",
            summary="Prevents unnecessary API surface growth or overexposed internals.",
            template="""
            ## Public Surface Control
            Keep the public surface area minimal.

            Only expose new types, members, endpoints, or settings when they are genuinely required by the use case.
            Prefer internal seams over expanding the public API without a clear consumer.
            """,
            tags=stable_tags("api-surface", "maintainability", "contracts"),
            prompt_types=["implementation", "refactor", "architecture", "security", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "architecture-spec", "security-hardening", "embedded-firmware-iteration"],
            phases=["architecture", "implementation", "verification"],
        ),
        component(
            key="feature-flag-rollout",
            name="Implementation: Feature Flag Rollout",
            group="implementation-execution",
            block_kind="Instruction",
            summary="Uses flags or staged switches when a risky behavior change should not land all at once.",
            template="""
            ## Feature Flag or Staged Rollout
            If {{target_change}} is risky, introduce it behind a feature flag, configuration switch, or staged default.

            Prefer:
            - additive wiring first,
            - test coverage on the new path,
            - flipping defaults only after validation.
            """,
            tags=stable_tags("feature-flag", "rollout", "risk"),
            prompt_types=["implementation", "migration", "performance", "security", "embedded"],
            blueprints=["feature-implementation", "performance-hardening", "security-hardening", "embedded-firmware-iteration"],
            phases=["implementation", "verification", "delivery"],
        ),
        component(
            key="run-build-after-each-slice",
            name="Implementation: Run Build After Each Slice",
            group="implementation-execution",
            block_kind="Testing",
            summary="Forces fast compile checks during implementation rather than only at the end.",
            template="""
            ## Build Checkpoints
            After each meaningful implementation slice:
            - run the fastest relevant build or compile command,
            - fix breakages before moving on,
            - keep the next slice starting from a working baseline.
            """,
            tags=stable_tags("build", "checkpoints", "implementation"),
            prompt_types=["implementation", "refactor", "bugfix", "migration", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration"],
            phases=["implementation", "verification"],
            recommended=True,
        ),
        component(
            key="document-changed-files",
            name="Implementation: Document Changed Files",
            group="implementation-execution",
            block_kind="Delivery",
            summary="Keeps a clear record of changed files grouped by responsibility.",
            template="""
            ## Changed Files Record
            Keep track of changed files grouped by responsibility area.

            The goal is not a changelog for its own sake. It is to make reviews, handoffs, and later debugging easier.
            """,
            tags=stable_tags("files-changed", "review", "handoff"),
            prompt_types=["implementation", "refactor", "bugfix", "migration", "embedded", "ui"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["implementation", "verification", "delivery"],
        ),
        component(
            key="manual-verification-steps",
            name="Implementation: Manual Verification Steps",
            group="implementation-execution",
            block_kind="Validation",
            summary="Requires the agent to leave reproducible manual verification instructions for the touched behavior.",
            template="""
            ## Manual Verification
            Provide manual verification steps for the changed behavior.

            The steps should state:
            - what to open or run,
            - what input or action to perform,
            - what the expected result should be,
            - what logs, UI state, or artifacts should appear.
            """,
            tags=stable_tags("manual-verification", "qa", "handoff"),
            prompt_types=["implementation", "bugfix", "ui", "embedded", "validation"],
            blueprints=["feature-implementation", "bugfix-with-regression-lock", "ui-ux-delivery", "embedded-firmware-iteration", "validation-audit"],
            phases=["verification", "delivery"],
        ),
        component(
            key="preserve-existing-contracts-and-data",
            name="Implementation: Preserve Existing Contracts and Data",
            group="implementation-execution",
            block_kind="Constraint",
            summary="Protects data shape, protocol shape, and existing consumers while implementing new behavior.",
            template="""
            ## Contract and Data Preservation
            While implementing {{target_change}}, preserve existing data and protocol expectations unless a planned migration says otherwise.

            Check:
            - serialized payloads,
            - storage format or schema assumptions,
            - API consumers,
            - user workflows that depend on the existing behavior.
            """,
            tags=stable_tags("contracts", "data", "compatibility"),
            prompt_types=["implementation", "refactor", "bugfix", "migration", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "embedded-firmware-iteration"],
            phases=["implementation", "verification"],
        ),
    ]


def build_validation_components() -> list[ComponentDefinition]:
    return [
        component(
            key="mandatory-unit-tests",
            name="Validation: Mandatory Unit Tests",
            group="validation-review",
            block_kind="Testing",
            summary="Forces unit tests for pure logic or tightly scoped behavior.",
            template="""
            ## Unit Tests
            Add or update unit tests for the logic touched by this work.

            The unit tests should:
            - target the smallest stable unit that covers the behavior,
            - lock in the failure mode being fixed or introduced,
            - remain deterministic and fast.
            """,
            tags=stable_tags("unit-tests", "testing", "logic"),
            prompt_types=["implementation", "refactor", "bugfix", "testing", "validation", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "test-strategy-and-automation", "validation-audit", "embedded-firmware-iteration"],
            phases=["implementation", "verification"],
            recommended=True,
        ),
        component(
            key="mandatory-integration-tests",
            name="Validation: Mandatory Integration Tests",
            group="validation-review",
            block_kind="Testing",
            summary="Requires integration or contract tests where boundaries matter.",
            template="""
            ## Integration or Contract Tests
            Add integration or contract tests wherever behavior crosses process, storage, or module boundaries.

            Use them to cover:
            - database behavior,
            - API contracts,
            - filesystem or network integration,
            - service composition that unit tests cannot prove.
            """,
            tags=stable_tags("integration-tests", "contracts", "boundaries"),
            prompt_types=["implementation", "migration", "testing", "validation", "security", "performance"],
            blueprints=["feature-implementation", "test-strategy-and-automation", "validation-audit", "security-hardening", "performance-hardening"],
            phases=["implementation", "verification"],
        ),
        component(
            key="mandatory-playwright-tests",
            name="Validation: Mandatory Playwright Tests",
            group="validation-review",
            block_kind="Testing",
            summary="Requires Playwright or browser-level tests for user-visible flows.",
            template="""
            ## UI End-to-End Tests
            Add or update Playwright coverage for any user-visible flow changed by this work.

            Focus on:
            - critical happy paths,
            - the regression being fixed,
            - the main state transition that proves the UI is wired correctly.
            """,
            tags=stable_tags("playwright", "ui-tests", "e2e"),
            prompt_types=["implementation", "bugfix", "ui", "testing", "validation"],
            blueprints=["feature-implementation", "bugfix-with-regression-lock", "test-strategy-and-automation", "ui-ux-delivery", "validation-audit"],
            phases=["implementation", "verification"],
        ),
        component(
            key="reviewer-findings-first",
            name="Validation: Reviewer Findings First",
            group="validation-review",
            block_kind="Validation",
            summary="Sets the output style for review prompts so findings come before summary.",
            template="""
            ## Review Output Style
            If this session is a review, present findings first.

            Order them by severity and include:
            - the risky behavior or flaw,
            - where it lives,
            - why it matters,
            - what evidence is missing or what should change.

            Keep the overall summary short and secondary.
            """,
            tags=stable_tags("review", "findings-first", "risk"),
            prompt_types=["review", "validation", "security", "performance"],
            blueprints=["senior-code-review", "validation-audit", "security-hardening", "performance-hardening"],
            phases=["verification", "delivery"],
            recommended=True,
        ),
        component(
            key="regression-proof-required",
            name="Validation: Regression Proof Required",
            group="validation-review",
            block_kind="Validation",
            summary="Requires explicit proof that the old bug or risk is now locked down.",
            template="""
            ## Regression Proof
            Provide regression proof for {{target_bug_or_risk}}.

            That proof can be:
            - an automated test,
            - a fixture comparison,
            - a screenshot, trace, or log,
            - a targeted manual reproduction with a documented result.

            Do not rely on "it should work now" reasoning.
            """,
            tags=stable_tags("regression", "evidence", "proof"),
            prompt_types=["bugfix", "refactor", "review", "validation", "testing"],
            blueprints=["bugfix-with-regression-lock", "safe-refactor", "validation-audit", "test-strategy-and-automation", "senior-code-review"],
            phases=["verification", "delivery"],
            recommended=True,
        ),
        component(
            key="architecture-validation-pass",
            name="Validation: Architecture Validation Pass",
            group="validation-review",
            block_kind="Validation",
            summary="Checks whether the implementation still respects the intended architecture.",
            template="""
            ## Architecture Validation
            Validate that the implementation still matches the intended architecture.

            Check:
            - module boundaries,
            - ownership of state and storage,
            - contract shape,
            - unwanted coupling or leakage between layers.
            """,
            tags=stable_tags("architecture-validation", "boundaries", "design"),
            prompt_types=["validation", "review", "architecture", "implementation"],
            blueprints=["validation-audit", "senior-code-review", "architecture-spec", "feature-implementation"],
            phases=["verification", "delivery"],
        ),
        component(
            key="performance-validation-pass",
            name="Validation: Performance Validation Pass",
            group="validation-review",
            block_kind="Validation",
            summary="Makes performance expectations explicit instead of optional.",
            template="""
            ## Performance Validation
            Validate the performance impact of {{target_change}}.

            Measure or reason explicitly about:
            - critical hot paths,
            - allocations or payload size,
            - latency or scheduling impact,
            - changes that could degrade mobile, browser, or embedded environments.
            """,
            tags=stable_tags("performance", "latency", "hot-path"),
            prompt_types=["performance", "implementation", "review", "embedded", "ui"],
            blueprints=["performance-hardening", "feature-implementation", "validation-audit", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["verification", "delivery"],
        ),
        component(
            key="security-validation-pass",
            name="Validation: Security Validation Pass",
            group="validation-review",
            block_kind="Security",
            summary="Requires a security-focused validation pass for sensitive changes.",
            template="""
            ## Security Validation
            Perform a security-focused review of {{target_change}}.

            Look for:
            - secret leakage,
            - unsafe storage or transport,
            - injection or validation gaps,
            - authz or authn regressions,
            - excessive trust in client-side inputs.
            """,
            tags=stable_tags("security", "review", "threats"),
            prompt_types=["security", "implementation", "review", "validation"],
            blueprints=["security-hardening", "feature-implementation", "senior-code-review", "validation-audit"],
            phases=["verification", "delivery"],
        ),
        component(
            key="accessibility-validation-pass",
            name="Validation: Accessibility Validation Pass",
            group="validation-review",
            block_kind="Validation",
            summary="Adds accessibility checks for changed UI surfaces.",
            template="""
            ## Accessibility Validation
            Validate accessibility for {{target_ui_flow}}.

            Check:
            - keyboard reachability,
            - focus order and visible focus,
            - labels and semantics,
            - fallback behavior when canvas or complex UI is involved.
            """,
            tags=stable_tags("accessibility", "a11y", "ui"),
            prompt_types=["ui", "implementation", "review", "validation"],
            blueprints=["ui-ux-delivery", "validation-audit", "feature-implementation", "senior-code-review"],
            phases=["verification", "delivery"],
        ),
        component(
            key="observability-validation-pass",
            name="Validation: Observability Validation Pass",
            group="validation-review",
            block_kind="Validation",
            summary="Makes logs, metrics, traces, or debug surfaces part of the quality bar when relevant.",
            template="""
            ## Observability Validation
            Confirm that {{target_change}} is observable enough to debug and support.

            Cover:
            - logs or traces for failure paths,
            - debug surfaces or diagnostics if needed,
            - redaction of sensitive values,
            - proof that the new behavior can be inspected when it fails.
            """,
            tags=stable_tags("observability", "logging", "diagnostics"),
            prompt_types=["implementation", "validation", "performance", "security", "embedded"],
            blueprints=["feature-implementation", "validation-audit", "performance-hardening", "security-hardening", "embedded-firmware-iteration"],
            phases=["verification", "delivery"],
        ),
        component(
            key="final-audit",
            name="Validation: Final Audit",
            group="validation-review",
            block_kind="Validation",
            summary="Forces a final audit after the main implementation is complete.",
            template="""
            ## Final Audit
            Perform a final audit before declaring completion.

            Confirm:
            - the success criteria are truly met,
            - the highest-risk regressions have proof,
            - the deliverables are present,
            - the remaining gaps are named honestly.
            """,
            tags=stable_tags("final-audit", "completion", "quality-gate"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["validation-audit", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "ui-ux-delivery", "embedded-firmware-iteration"],
            phases=["verification", "delivery"],
            recommended=True,
        ),
        component(
            key="evidence-output-required",
            name="Validation: Evidence Output Required",
            group="validation-review",
            block_kind="Delivery",
            summary="Requires exact commands, proof artifacts, or scenario outputs in the final response.",
            template="""
            ## Evidence Output
            In the final response, include the evidence that supports the result:
            - exact commands executed,
            - test or build outcome summary,
            - screenshots, traces, or logs if relevant,
            - what could not be verified in this environment.
            """,
            tags=stable_tags("evidence", "commands", "proof"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["validation-audit", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "test-strategy-and-automation", "embedded-firmware-iteration"],
            phases=["verification", "delivery"],
            recommended=True,
        ),
    ]


def build_output_components() -> list[ComponentDefinition]:
    return [
        component(
            key="output-scope-summary",
            name="Output: Scope Summary",
            group="output-handoff",
            block_kind="Delivery",
            summary="Standardizes a brief restatement of what the session covered.",
            template="""
            ## Scope Summary
            Start the close-out with a short scope summary covering:
            - what the session was meant to accomplish,
            - what area of the repo or system it touched,
            - whether the result is complete, partial, or intentionally phase-gated.
            """,
            tags=stable_tags("output", "summary", "scope"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "senior-code-review", "validation-audit", "ui-ux-delivery", "embedded-firmware-iteration"],
            phases=["delivery"],
            recommended=True,
        ),
        component(
            key="output-implementation-plan",
            name="Output: Implementation Plan",
            group="output-handoff",
            block_kind="Delivery",
            summary="Standardizes how a plan should be presented when the session is planning-oriented.",
            template="""
            ## Implementation Plan Output
            When the session is planning-oriented, present the implementation plan as a flat, ordered sequence.

            Each item should be clear enough that another agent can execute it without rediscovering the intent.
            """,
            tags=stable_tags("output", "plan", "handoff"),
            prompt_types=["architecture", "plan", "migration"],
            blueprints=["architecture-spec", "implementation-plan"],
            phases=["delivery"],
        ),
        component(
            key="output-files-changed",
            name="Output: Files Changed",
            group="output-handoff",
            block_kind="Delivery",
            summary="Requires changed files to be summarized clearly in the final response.",
            template="""
            ## Files Changed
            Include the changed files or artifacts grouped by responsibility area.

            Keep the list readable and focused on the signal, not a raw dump of every low-value file.
            """,
            tags=stable_tags("output", "files", "changes"),
            prompt_types=["implementation", "refactor", "bugfix", "migration", "ui", "embedded"],
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "ui-ux-delivery", "embedded-firmware-iteration"],
            phases=["delivery"],
        ),
        component(
            key="output-commands-executed",
            name="Output: Commands Executed",
            group="output-handoff",
            block_kind="Delivery",
            summary="Requires the final response to name the important verification commands run.",
            template="""
            ## Commands Executed
            Include the important commands executed for build, tests, migrations, or tooling.

            If a command failed or could not run, state that clearly and explain why.
            """,
            tags=stable_tags("output", "commands", "verification"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "test-strategy-and-automation", "validation-audit", "embedded-firmware-iteration"],
            phases=["delivery"],
            recommended=True,
        ),
        component(
            key="output-completion-summary",
            name="Output: Completion Summary",
            group="output-handoff",
            block_kind="Delivery",
            summary="Standardizes the final completion or non-completion statement.",
            template="""
            ## Completion Summary
            State clearly:
            - what is complete,
            - what is not complete,
            - what the highest remaining risk is.

            Avoid vague phrases such as "mostly done" without specifics.
            """,
            tags=stable_tags("output", "completion", "status"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "ui-ux-delivery", "embedded-firmware-iteration"],
            phases=["delivery"],
            recommended=True,
        ),
        component(
            key="output-remaining-risks-and-next-steps",
            name="Output: Remaining Risks and Next Steps",
            group="output-handoff",
            block_kind="Delivery",
            summary="Keeps residual risk and next actions visible at handoff.",
            template="""
            ## Remaining Risks and Next Steps
            Close with:
            - remaining risks,
            - unresolved assumptions,
            - the highest-value next action.

            If there is a specific next prompt or next agent, name it directly.
            """,
            tags=stable_tags("output", "risks", "next-steps"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "ui-ux-delivery", "embedded-firmware-iteration"],
            phases=["delivery"],
            recommended=True,
        ),
        component(
            key="handoff-to-next-agent",
            name="Output: Handoff to Next Agent",
            group="output-handoff",
            block_kind="Delivery",
            summary="Creates a compact handoff format for the next agent in a chain.",
            template="""
            ## Handoff
            When another agent will continue the work, include:
            - the current state,
            - artifacts or files they must read first,
            - the next decision they must make,
            - the constraints and risks they inherit.
            """,
            tags=stable_tags("handoff", "next-agent", "continuity"),
            prompt_types=["architecture", "plan", "implementation", "review", "validation", "migration", "embedded"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "validation-audit", "embedded-firmware-iteration"],
            phases=["delivery"],
        ),
        component(
            key="artifact-list-and-links",
            name="Output: Artifact List and Links",
            group="output-handoff",
            block_kind="Delivery",
            summary="Keeps docs, spreadsheets, diagrams, screenshots, or logs visible as explicit outputs.",
            template="""
            ## Artifacts
            List the generated or updated artifacts:
            - docs,
            - diagrams,
            - spreadsheets,
            - screenshots or videos,
            - logs or reports.

            Use stable file paths and call out anything that should be reviewed manually.
            """,
            tags=stable_tags("artifacts", "docs", "evidence"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "validation-audit", "ui-ux-delivery", "embedded-firmware-iteration"],
            phases=["delivery"],
        ),
    ]


def build_stack_components() -> list[ComponentDefinition]:
    return [
        component(
            key="stack-dotnet-solution",
            name="Stack: .NET Solution",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies .NET solution-level guidance for build, tests, layering, and project wiring.",
            template="""
            ## .NET Guidance
            Treat the real solution or project graph as authoritative.

            Requirements:
            - use the correct solution or project entry point instead of guessing,
            - preserve dependency injection, nullable, analyzers, and test conventions already used in the repo,
            - keep domain logic out of thin UI or host layers,
            - if you add a project or contract, wire references, configuration, and tests in the same session.

            Primary commands:
            - build: {{dotnet_build_command}}
            - test: {{dotnet_test_command}}
            """,
            tags=stable_tags(".net", "solution", "build", "tests"),
            prompt_types=["architecture", "plan", "implementation", "refactor", "bugfix", "testing", "validation", "performance", "security", "migration"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "validation-audit", "performance-hardening", "security-hardening"],
            phases=["discovery", "planning", "implementation", "verification"],
            stack_tags=[".net"],
            recommended=True,
        ),
        component(
            key="stack-blazor-webapp",
            name="Stack: Blazor",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies Blazor-specific guidance for render modes, components, and JS interop boundaries.",
            template="""
            ## Blazor Guidance
            For Blazor work:
            - keep business logic out of page-only code,
            - respect render mode boundaries and lifecycle realities,
            - keep component state, services, and JS interop responsibilities explicit,
            - preserve the existing component system and routing conventions.

            If a behavior depends on browser state or JS interop, document the contract on both sides of the boundary.
            """,
            tags=stable_tags("blazor", "components", "js-interop"),
            prompt_types=["architecture", "implementation", "refactor", "bugfix", "ui", "testing", "validation"],
            blueprints=["architecture-spec", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "ui-ux-delivery", "validation-audit"],
            phases=["architecture", "planning", "implementation", "verification"],
            stack_tags=["blazor", ".net"],
            recommended=True,
        ),
        component(
            key="stack-html-js-css",
            name="Stack: HTML/JS/CSS",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies browser-first frontend guidance when working directly with HTML, JavaScript, and CSS.",
            template="""
            ## HTML, JavaScript, and CSS Guidance
            For direct frontend work:
            - respect the existing asset pipeline and project structure,
            - keep behavior, layout, and styling responsibilities clear,
            - test edge states such as small screens, long content, and async failures,
            - avoid adding heavy dependencies unless the prompt explicitly allows them.
            """,
            tags=stable_tags("html", "javascript", "css", "frontend"),
            prompt_types=["architecture", "implementation", "bugfix", "ui", "testing", "validation"],
            blueprints=["feature-implementation", "bugfix-with-regression-lock", "ui-ux-delivery", "validation-audit"],
            phases=["planning", "implementation", "verification"],
            stack_tags=["html", "javascript", "css"],
            recommended=True,
        ),
        component(
            key="stack-tailwind-css",
            name="Stack: Tailwind CSS",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies Tailwind-specific guidance for utility usage, design consistency, and component reuse.",
            template="""
            ## Tailwind Guidance
            When the repo uses Tailwind:
            - preserve the existing design tokens and utility conventions,
            - extract reusable patterns when class lists become structural rather than incidental,
            - avoid style drift by matching spacing, typography, and interaction patterns already in use,
            - keep accessibility and responsive behavior visible in the implementation.
            """,
            tags=stable_tags("tailwind", "css", "design-system"),
            prompt_types=["implementation", "ui", "review", "validation"],
            blueprints=["ui-ux-delivery", "feature-implementation", "validation-audit"],
            phases=["planning", "implementation", "verification"],
            stack_tags=["tailwind", "css"],
        ),
        component(
            key="stack-php-webapp",
            name="Stack: PHP Web App",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies PHP web app guidance for legacy code, mixed rendering, and safer incremental change.",
            template="""
            ## PHP Web App Guidance
            For PHP-based web apps:
            - inspect the real runtime structure before assuming framework boundaries,
            - preserve working server-side rendering, routing, and data flow unless the prompt calls for migration,
            - keep new JavaScript and CSS changes compatible with the existing PHP entry points,
            - prefer incremental modernization over hidden framework rewrites.
            """,
            tags=stable_tags("php", "webapp", "legacy-modernization"),
            prompt_types=["architecture", "implementation", "refactor", "bugfix", "migration", "ui"],
            blueprints=["architecture-spec", "feature-implementation", "safe-refactor", "bugfix-with-regression-lock", "ui-ux-delivery"],
            phases=["discovery", "planning", "implementation", "verification"],
            stack_tags=["php"],
            recommended=True,
        ),
        component(
            key="stack-playwright-mcp",
            name="Stack: Playwright MCP",
            group="stack-profiles",
            block_kind="Testing",
            summary="Applies Playwright MCP guidance for real-browser automation, evidence capture, and UI debugging.",
            template="""
            ## Playwright MCP Guidance
            When browser automation is useful:
            - use Playwright MCP or the closest real-browser path instead of shallow DOM reasoning,
            - capture screenshots, traces, or recordings when they materially improve proof,
            - validate key flows end to end instead of only checking that elements render.
            """,
            tags=stable_tags("playwright", "mcp", "browser-automation", "testing"),
            prompt_types=["testing", "validation", "ui", "implementation", "bugfix"],
            blueprints=["test-strategy-and-automation", "validation-audit", "ui-ux-delivery", "feature-implementation", "bugfix-with-regression-lock"],
            phases=["implementation", "verification", "delivery"],
            stack_tags=["playwright", "mcp"],
        ),
        component(
            key="stack-canvas-html-js",
            name="Stack: Canvas in HTML/JS",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies canvas-specific guidance for rendering, hit testing, DOM boundaries, and performance.",
            template="""
            ## Canvas Guidance
            For HTML5 canvas work:
            - keep the actual interactive surface canvas-first if that is the product intent,
            - use DOM only where it is clearly the correct tool,
            - make hit testing, coordinate transforms, resize behavior, and redraw costs explicit,
            - reuse existing canvas primitives before inventing a parallel rendering stack.
            """,
            tags=stable_tags("canvas", "html5", "javascript", "rendering"),
            prompt_types=["architecture", "implementation", "bugfix", "ui", "performance", "validation"],
            blueprints=["architecture-spec", "feature-implementation", "bugfix-with-regression-lock", "ui-ux-delivery", "performance-hardening", "validation-audit"],
            phases=["architecture", "planning", "implementation", "verification"],
            stack_tags=["canvas", "html", "javascript"],
            recommended=True,
        ),
        component(
            key="stack-postgresql",
            name="Stack: PostgreSQL",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies PostgreSQL-specific guidance for schema, queries, indexes, and migrations.",
            template="""
            ## PostgreSQL Guidance
            For PostgreSQL-backed work:
            - design schemas and indexes around actual query paths,
            - respect transaction and migration safety,
            - be explicit about JSON, array, text search, or extension usage,
            - confirm behavior in explicit in-memory test overrides where relevant.
            """,
            tags=stable_tags("postgresql", "database", "schema"),
            prompt_types=["architecture", "plan", "implementation", "migration", "performance", "validation"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "performance-hardening", "validation-audit"],
            phases=["architecture", "planning", "implementation", "verification"],
            stack_tags=["postgresql", "database"],
        ),
        component(
            key="stack-efcore",
            name="Stack: EF Core",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies EF Core guidance for DbContext usage, migrations, query shape, and testability.",
            template="""
            ## EF Core Guidance
            For EF Core work:
            - keep `DbContext` lifetime and ownership clear,
            - shape entities and configurations explicitly,
            - watch for N+1 patterns, tracking mistakes, and provider-specific drift,
            - pair model changes with migrations, tests, and rollback notes.
            """,
            tags=stable_tags("efcore", "orm", "database"),
            prompt_types=["architecture", "plan", "implementation", "refactor", "migration", "performance", "validation"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "safe-refactor", "performance-hardening", "validation-audit"],
            phases=["architecture", "planning", "implementation", "verification"],
            stack_tags=["efcore", ".net", "database"],
            recommended=True,
        ),
        component(
            key="stack-arduino-firmware",
            name="Stack: Arduino Firmware",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies embedded firmware guidance for limited resources, state machines, and hardware reliability.",
            template="""
            ## Arduino Firmware Guidance
            For Arduino-class firmware work:
            - treat memory, timing, and pin ownership as hard constraints,
            - prefer deterministic state machines over implicit control flow,
            - minimize heap churn in hot paths,
            - make calibration, thresholds, and protocol behavior observable and testable.
            """,
            tags=stable_tags("arduino", "firmware", "embedded", "realtime"),
            prompt_types=["embedded", "implementation", "bugfix", "performance", "validation"],
            blueprints=["embedded-firmware-iteration", "feature-implementation", "validation-audit", "performance-hardening"],
            phases=["discovery", "planning", "implementation", "verification"],
            stack_tags=["arduino", "embedded"],
            recommended=True,
        ),
        component(
            key="stack-midi-audio",
            name="Stack: MIDI and Audio",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies timing and signal-path guidance for MIDI, quantization, and audio scheduling work.",
            template="""
            ## MIDI and Audio Guidance
            For MIDI or audio work:
            - make time bases explicit,
            - protect event ordering and timing accuracy,
            - use fixtures or captured traces when live hardware input is unavailable,
            - avoid hand-wavy assumptions around quantization, buffering, or scheduling.
            """,
            tags=stable_tags("midi", "audio", "timing", "realtime"),
            prompt_types=["embedded", "implementation", "bugfix", "performance", "validation", "ui"],
            blueprints=["embedded-firmware-iteration", "feature-implementation", "bugfix-with-regression-lock", "validation-audit", "performance-hardening"],
            phases=["architecture", "planning", "implementation", "verification"],
            stack_tags=["midi", "audio"],
            recommended=True,
        ),
        component(
            key="stack-m5stack",
            name="Stack: M5Stack",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies M5Stack-specific guidance for pins, PMU, display, sensors, and board-level constraints.",
            template="""
            ## M5Stack Guidance
            For M5Stack or M5Stick-class work:
            - verify the exact board model and pin assignments,
            - be deliberate about PMU, battery telemetry, and peripheral initialization,
            - do not repurpose reserved pins without proving the impact,
            - connect firmware behavior to any host-side telemetry or UI surfaces that depend on it.
            """,
            tags=stable_tags("m5stack", "hardware", "embedded"),
            prompt_types=["embedded", "implementation", "bugfix", "review", "validation"],
            blueprints=["embedded-firmware-iteration", "validation-audit", "feature-implementation"],
            phases=["discovery", "planning", "implementation", "verification"],
            stack_tags=["m5stack", "embedded"],
        ),
        component(
            key="stack-offline-first-sync",
            name="Stack: Offline-First Sync",
            group="stack-profiles",
            block_kind="Instruction",
            summary="Applies offline-first guidance for local state, outbox sync, conflicts, and low-connectivity behavior.",
            template="""
            ## Offline-First Guidance
            For offline-first or sync-heavy work:
            - keep local state ownership explicit,
            - model outbox, retries, idempotency, and conflict handling deliberately,
            - make online and offline states visible in the UX,
            - prove behavior across reconnect and partial failure paths.
            """,
            tags=stable_tags("offline-first", "sync", "local-state"),
            prompt_types=["architecture", "plan", "implementation", "migration", "validation", "ui"],
            blueprints=["architecture-spec", "implementation-plan", "feature-implementation", "validation-audit", "ui-ux-delivery"],
            phases=["architecture", "planning", "implementation", "verification"],
            stack_tags=["offline-first", "sync"],
            recommended=True,
        ),
    ]


def build_toolbox_components() -> list[ComponentDefinition]:
    return [
        component(
            key="toolbox-run-unit-tests-docker",
            name="Toolbox: Run Unit Tests in Docker",
            group="toolbox-snippets",
            block_kind="Testing",
            summary="Short insert that forces unit tests to run in Docker with cache-aware resource usage.",
            template="""
            ## Unit Tests in Docker
            You must run the unit test suite inside Docker before declaring this work done.

            Requirements:
            - use {{docker_compose_file_or_dockerfile}} if available, otherwise create the smallest viable temporary test container,
            - reuse package, image, and layer caches whenever possible to reduce network transfer and save mobile data,
            - print the exact Docker command, target test projects, and result summary,
            - if Docker validation is blocked, say so clearly and fall back to the nearest reproducible local command without pretending Docker passed.
            """,
            tags=stable_tags("toolbox", "docker", "unit-tests", "mobile-data"),
            prompt_types=["implementation", "bugfix", "refactor", "testing", "validation"],
            blueprints=["feature-implementation", "bugfix-with-regression-lock", "safe-refactor", "test-strategy-and-automation", "validation-audit"],
            phases=["verification", "delivery"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-run-integration-tests-docker",
            name="Toolbox: Run Integration Tests in Docker",
            group="toolbox-snippets",
            block_kind="Testing",
            summary="Short insert that forces integration or API tests to run in Docker.",
            template="""
            ## Integration Tests in Docker
            Run the integration or contract test suite inside Docker.

            Requirements:
            - start only the dependencies the tests truly need,
            - use persistent caches and volumes where safe to reduce repeated downloads,
            - record the exact command and the services started,
            - if external dependencies make Docker validation impossible here, explain the blocker and the closest fallback.
            """,
            tags=stable_tags("toolbox", "docker", "integration-tests"),
            prompt_types=["implementation", "migration", "testing", "validation", "security", "performance"],
            blueprints=["feature-implementation", "test-strategy-and-automation", "validation-audit", "security-hardening", "performance-hardening"],
            phases=["verification", "delivery"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-run-ui-tests-docker",
            name="Toolbox: Run UI Tests in Docker",
            group="toolbox-snippets",
            block_kind="Testing",
            summary="Short insert that forces Playwright or browser UI tests to run in Docker.",
            template="""
            ## UI Tests in Docker
            Run the UI or Playwright suite inside Docker.

            Requirements:
            - install only the browsers and system dependencies actually needed,
            - capture screenshots, traces, or reports for failures,
            - reuse browser and package caches where possible to save bandwidth,
            - include the exact command and artifact locations in the final output.
            """,
            tags=stable_tags("toolbox", "docker", "playwright", "ui-tests", "mobile-data"),
            prompt_types=["implementation", "bugfix", "ui", "testing", "validation"],
            blueprints=["feature-implementation", "bugfix-with-regression-lock", "ui-ux-delivery", "test-strategy-and-automation", "validation-audit"],
            phases=["verification", "delivery"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-use-playwright-mcp-now",
            name="Toolbox: Use Playwright MCP Now",
            group="toolbox-snippets",
            block_kind="Testing",
            summary="Directs the agent to validate through Playwright MCP rather than static reasoning only.",
            template="""
            ## Use Playwright MCP
            Validate this UI flow with Playwright MCP or the closest real browser automation path available.

            Do not rely only on reading the code when:
            - interaction timing matters,
            - canvas or drag/drop behavior matters,
            - responsive behavior matters,
            - the reported bug is visual or browser-state dependent.
            """,
            tags=stable_tags("toolbox", "playwright", "mcp", "browser"),
            prompt_types=["ui", "bugfix", "testing", "validation"],
            blueprints=["ui-ux-delivery", "bugfix-with-regression-lock", "test-strategy-and-automation", "validation-audit"],
            phases=["verification"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-capture-browser-artifacts",
            name="Toolbox: Capture Browser Artifacts",
            group="toolbox-snippets",
            block_kind="Delivery",
            summary="Adds screenshot, trace, log, or video capture expectations for browser validation.",
            template="""
            ## Capture Browser Artifacts
            Collect browser evidence for the key scenarios:
            - screenshots for final or failing states,
            - traces when interaction or timing bugs matter,
            - logs or network evidence when data flow matters,
            - video only when it materially improves diagnosis.

            Include the saved artifact paths in the final output.
            """,
            tags=stable_tags("toolbox", "screenshots", "trace", "artifacts"),
            prompt_types=["ui", "testing", "validation", "bugfix"],
            blueprints=["ui-ux-delivery", "test-strategy-and-automation", "validation-audit", "bugfix-with-regression-lock"],
            phases=["verification", "delivery"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-db-migration-dry-run",
            name="Toolbox: Database Migration Dry Run",
            group="toolbox-snippets",
            block_kind="Testing",
            summary="Adds a migration dry run before data-layer work is considered complete.",
            template="""
            ## Database Migration Dry Run
            Before declaring the data-layer work done:
            - generate or apply the migration in a safe non-production environment,
            - validate upgrade and downgrade behavior if supported,
            - record the exact commands and any warnings,
            - call out any provider-specific caveats.
            """,
            tags=stable_tags("toolbox", "database", "migration", "dry-run"),
            prompt_types=["implementation", "migration", "validation", "performance", "security"],
            blueprints=["feature-implementation", "implementation-plan", "validation-audit", "performance-hardening", "security-hardening"],
            phases=["verification", "delivery"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-cross-db-compat-check",
            name="Toolbox: Persistence Compatibility Check",
            group="toolbox-snippets",
            block_kind="Testing",
            summary="Adds compatibility checks across PostgreSQL and explicit test-provider overrides.",
            template="""
            ## Persistence Compatibility
            Validate this data-layer change across the supported runtime and test providers.

            At minimum:
            - note which providers were tested,
            - identify provider-specific behavior or skipped coverage,
            - avoid assuming in-memory test behavior proves PostgreSQL translation or concurrency.
            """,
            tags=stable_tags("toolbox", "database", "compatibility", "postgresql"),
            prompt_types=["implementation", "migration", "testing", "validation"],
            blueprints=["feature-implementation", "implementation-plan", "test-strategy-and-automation", "validation-audit"],
            phases=["verification", "delivery"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-cache-downloads-mobile-data",
            name="Toolbox: Cache Downloads to Save Mobile Data",
            group="toolbox-snippets",
            block_kind="Constraint",
            summary="Adds a reusable instruction to reuse caches and avoid wasteful downloads.",
            template="""
            ## Cache-Aware Resource Usage
            Be careful with network-heavy setup steps.

            Reuse:
            - dependency caches,
            - Docker layers,
            - browser caches,
            - package manager caches,
            whenever it is safe to do so.

            The goal is to save bandwidth and mobile data without compromising reproducibility.
            """,
            tags=stable_tags("toolbox", "caching", "mobile-data", "bandwidth"),
            prompt_types=ALL_PROMPT_TYPES,
            blueprints=["feature-implementation", "test-strategy-and-automation", "validation-audit", "embedded-firmware-iteration"],
            phases=["implementation", "verification"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-generate-fixtures-and-seed-data",
            name="Toolbox: Generate Fixtures and Seed Data",
            group="toolbox-snippets",
            block_kind="Testing",
            summary="Adds a reusable instruction to create focused fixtures or seed data for proof.",
            template="""
            ## Fixtures and Seed Data
            Create focused fixtures or seed data for the scenarios being changed.

            Prefer:
            - the smallest data that still reproduces the behavior,
            - deterministic values,
            - fixtures that can be reused by unit, integration, or UI tests.
            """,
            tags=stable_tags("toolbox", "fixtures", "seed-data", "testing"),
            prompt_types=["implementation", "bugfix", "testing", "validation", "embedded", "ui"],
            blueprints=["feature-implementation", "bugfix-with-regression-lock", "test-strategy-and-automation", "validation-audit", "embedded-firmware-iteration", "ui-ux-delivery"],
            phases=["planning", "implementation", "verification"],
            toolbox_eligible=True,
        ),
        component(
            key="toolbox-create-manual-qa-checklist",
            name="Toolbox: Create Manual QA Checklist",
            group="toolbox-snippets",
            block_kind="Validation",
            summary="Adds a short manual QA list when automated coverage is not enough.",
            template="""
            ## Manual QA Checklist
            Create a short manual QA checklist for the changed flow.

            Include:
            - the scenario to run,
            - expected result,
            - failure clues to watch for,
            - any environment prerequisites.
            """,
            tags=stable_tags("toolbox", "manual-qa", "checklist"),
            prompt_types=["testing", "validation", "ui", "embedded", "implementation"],
            blueprints=["test-strategy-and-automation", "validation-audit", "ui-ux-delivery", "embedded-firmware-iteration", "feature-implementation"],
            phases=["verification", "delivery"],
            toolbox_eligible=True,
        ),
    ]


def build_components() -> list[ComponentDefinition]:
    components = []
    builders = [
        build_role_components,
        build_mission_scope_components,
        build_context_discovery_components,
        build_guardrail_components,
        build_workflow_components,
        build_architecture_analysis_components,
        build_planning_components,
        build_implementation_components,
        build_validation_components,
        build_output_components,
        build_stack_components,
        build_toolbox_components,
    ]

    for builder in builders:
        components.extend(builder())

    keys = [component.key for component in components]
    if len(keys) != len(set(keys)):
        duplicates = [key for key, count in Counter(keys).items() if count > 1]
        raise ValueError(f"Duplicate component keys: {duplicates}")
    return components


def build_flows() -> list[FlowDefinition]:
    return [
        FlowDefinition(
            key="architecture-review-plan-implement-validate",
            name="Architecture -> Review -> Plan -> Implement -> Validate",
            summary="Canonical multi-agent flow for new features or major changes.",
            prompt_types=["architecture", "plan", "implementation", "review", "validation"],
            block_keys=[
                "mission-exact-goal",
                "scope-in-scope-items",
                "scope-out-of-scope-items",
                "success-criteria",
                "repo-map-confirmation",
                "required-reading-list",
                "current-state-audit",
                "non-negotiable-rules",
                "preserve-architecture-boundaries",
                "small-verifiable-increments",
                "safe-ambiguity-handling",
                "multi-agent-chain",
                "sequential-phases",
                "required-phase-output-format",
                "architecture-blueprint",
                "gap-analysis",
                "risk-register",
                "implementation-plan-step-by-step",
                "dependency-ordering",
                "test-plan-matrix",
                "definition-of-done",
                "implement-in-small-slices",
                "additive-refactor-first",
                "mandatory-unit-tests",
                "mandatory-integration-tests",
                "regression-proof-required",
                "final-audit",
                "evidence-output-required",
                "output-completion-summary",
                "output-remaining-risks-and-next-steps",
            ],
            agent_sequence=[
                FlowAgentStep(1, "role-architecture-lead", "architecture-spec", "architecture", "Produce the implementation-ready design and risk model.", ["mission-exact-goal", "scope-in-scope-items", "repo-map-confirmation", "required-reading-list", "current-state-audit", "architecture-blueprint", "gap-analysis", "risk-register"]),
                FlowAgentStep(2, "role-senior-reviewer", "senior-code-review", "verification", "Challenge the architecture and expose hidden risks or missing proof.", ["success-criteria", "reviewer-findings-first", "architecture-validation-pass", "security-validation-pass", "performance-validation-pass"]),
                FlowAgentStep(3, "role-implementation-planner", "implementation-plan", "planning", "Convert the approved architecture into milestones, files, tests, and checklists.", ["implementation-plan-step-by-step", "dependency-ordering", "files-and-modules-likely-involved", "test-plan-matrix", "definition-of-done", "acceptance-checklist"]),
                FlowAgentStep(4, "role-implementation-lead", "feature-implementation", "implementation", "Implement the work in stable slices with immediate verification.", ["implement-in-small-slices", "additive-refactor-first", "run-build-after-each-slice", "mandatory-unit-tests", "mandatory-integration-tests", "document-changed-files"]),
                FlowAgentStep(5, "role-test-validation-lead", "validation-audit", "delivery", "Gather final evidence, challenge gaps, and close out honestly.", ["regression-proof-required", "final-audit", "evidence-output-required", "output-commands-executed", "output-completion-summary", "output-remaining-risks-and-next-steps"]),
            ],
        ),
        FlowDefinition(
            key="audit-plan-refactor-review",
            name="Audit -> Plan -> Refactor -> Review",
            summary="Focused flow for structural cleanup with regression control.",
            prompt_types=["audit", "plan", "refactor", "review", "validation"],
            block_keys=["mission-exact-goal", "current-state-audit", "file-touch-plan", "preserve-backward-compatibility", "small-verifiable-increments", "multi-agent-chain", "required-phase-output-format", "implementation-plan-step-by-step", "files-and-modules-likely-involved", "rollback-and-recovery-plan", "implement-in-small-slices", "additive-refactor-first", "mandatory-unit-tests", "regression-proof-required", "reviewer-findings-first", "final-audit", "output-completion-summary", "output-remaining-risks-and-next-steps"],
            agent_sequence=[
                FlowAgentStep(1, "role-refactor-specialist", "repository-audit", "discovery", "Audit current structure and lock risky behavior.", ["current-state-audit", "file-touch-plan", "regression-proof-required"]),
                FlowAgentStep(2, "role-implementation-planner", "implementation-plan", "planning", "Define the safest refactor sequence.", ["implementation-plan-step-by-step", "files-and-modules-likely-involved", "rollback-and-recovery-plan"]),
                FlowAgentStep(3, "role-refactor-specialist", "safe-refactor", "implementation", "Apply the structural cleanup without drift.", ["implement-in-small-slices", "additive-refactor-first", "mandatory-unit-tests"]),
                FlowAgentStep(4, "role-senior-reviewer", "validation-audit", "delivery", "Audit the final refactor for regressions and risk.", ["reviewer-findings-first", "final-audit"]),
            ],
        ),
        FlowDefinition(
            key="bugfix-regression-proof",
            name="Bugfix -> Regression Proof -> Close",
            summary="Fast but disciplined flow for fixing a bug and proving it stays fixed.",
            prompt_types=["bugfix", "testing", "validation", "review"],
            block_keys=["mission-exact-goal", "success-criteria", "current-state-audit", "safe-ambiguity-handling", "implement-in-small-slices", "mandatory-unit-tests", "mandatory-playwright-tests", "regression-proof-required", "manual-verification-steps", "evidence-output-required", "output-completion-summary"],
            agent_sequence=[
                FlowAgentStep(1, "role-implementation-lead", "bugfix-with-regression-lock", "implementation", "Reproduce and fix the bug in the smallest safe slice.", ["current-state-audit", "implement-in-small-slices", "mandatory-unit-tests"]),
                FlowAgentStep(2, "role-test-validation-lead", "validation-audit", "verification", "Gather regression proof and UI proof where needed.", ["mandatory-playwright-tests", "regression-proof-required", "manual-verification-steps"]),
                FlowAgentStep(3, "role-senior-reviewer", "senior-code-review", "delivery", "Review the fix for hidden regressions or missing evidence.", ["reviewer-findings-first", "evidence-output-required", "output-completion-summary"]),
            ],
        ),
        FlowDefinition(
            key="ui-canvas-feature-delivery",
            name="UI and Canvas Feature Delivery",
            summary="Flow for complex interactive UI work where design, browser behavior, and evidence all matter.",
            prompt_types=["ui", "architecture", "implementation", "testing", "validation"],
            block_keys=["mission-exact-goal", "success-criteria", "repo-map-confirmation", "required-reading-list", "current-state-audit", "small-verifiable-increments", "multi-agent-chain", "required-phase-output-format", "ux-surface-map", "architecture-blueprint", "implementation-plan-step-by-step", "implement-in-small-slices", "no-placeholder-ui", "mandatory-playwright-tests", "accessibility-validation-pass", "performance-validation-pass", "final-audit", "output-completion-summary", "output-remaining-risks-and-next-steps"],
            agent_sequence=[
                FlowAgentStep(1, "role-ui-ux-engineer", "ui-ux-delivery", "architecture", "Map the flow and define the interaction architecture.", ["ux-surface-map", "architecture-blueprint", "success-criteria"]),
                FlowAgentStep(2, "role-implementation-planner", "implementation-plan", "planning", "Sequence UI, state, and test work into milestones.", ["implementation-plan-step-by-step", "test-plan-matrix"]),
                FlowAgentStep(3, "role-implementation-lead", "feature-implementation", "implementation", "Build the feature in slices using the real component and canvas stack.", ["implement-in-small-slices", "no-placeholder-ui", "mandatory-playwright-tests"]),
                FlowAgentStep(4, "role-test-validation-lead", "validation-audit", "delivery", "Validate browser behavior, accessibility, and performance.", ["accessibility-validation-pass", "performance-validation-pass", "final-audit"]),
            ],
        ),
        FlowDefinition(
            key="fullstack-offline-feature",
            name="Full-Stack Offline-First Feature",
            summary="Flow for features that span local state, sync, API, DB, and UI.",
            prompt_types=["architecture", "plan", "implementation", "migration", "validation"],
            block_keys=["mission-exact-goal", "mission-business-context", "scope-in-scope-items", "success-criteria", "repo-map-confirmation", "current-state-audit", "dependency-inventory", "preserve-backward-compatibility", "safe-ambiguity-handling", "multi-agent-chain", "required-phase-output-format", "architecture-blueprint", "data-model-and-migration-design", "api-contract-design", "risk-register", "implementation-plan-step-by-step", "test-plan-matrix", "definition-of-done", "implement-in-small-slices", "mandatory-unit-tests", "mandatory-integration-tests", "final-audit", "output-completion-summary", "output-remaining-risks-and-next-steps"],
            agent_sequence=[
                FlowAgentStep(1, "role-architecture-lead", "architecture-spec", "architecture", "Define storage, sync, API, and UI boundaries.", ["architecture-blueprint", "data-model-and-migration-design", "api-contract-design", "risk-register"]),
                FlowAgentStep(2, "role-senior-reviewer", "senior-code-review", "verification", "Challenge conflict handling, compatibility, and risk assumptions.", ["reviewer-findings-first", "architecture-validation-pass", "security-validation-pass"]),
                FlowAgentStep(3, "role-implementation-planner", "implementation-plan", "planning", "Translate the design into ordered full-stack milestones.", ["implementation-plan-step-by-step", "test-plan-matrix", "definition-of-done"]),
                FlowAgentStep(4, "role-implementation-lead", "feature-implementation", "implementation", "Implement the backend, sync, and UI slices with proof.", ["implement-in-small-slices", "mandatory-unit-tests", "mandatory-integration-tests"]),
                FlowAgentStep(5, "role-test-validation-lead", "validation-audit", "delivery", "Verify online, offline, and reconnect behavior with evidence.", ["final-audit", "evidence-output-required", "manual-verification-steps"]),
            ],
        ),
        FlowDefinition(
            key="data-layer-change-crossdb",
            name="Data-Layer Change with Cross-DB Proof",
            summary="Flow for EF Core, PostgreSQL, and migration-heavy work.",
            prompt_types=["architecture", "plan", "implementation", "migration", "testing", "validation"],
            block_keys=["mission-exact-goal", "current-state-audit", "dependency-inventory", "data-model-and-migration-design", "risk-register", "implementation-plan-step-by-step", "rollback-and-recovery-plan", "implement-in-small-slices", "mandatory-unit-tests", "mandatory-integration-tests", "final-audit"],
            agent_sequence=[
                FlowAgentStep(1, "role-architecture-lead", "architecture-spec", "architecture", "Design schema and migration changes with rollback safety.", ["data-model-and-migration-design", "risk-register"]),
                FlowAgentStep(2, "role-implementation-planner", "implementation-plan", "planning", "Sequence model changes, migrations, and tests.", ["implementation-plan-step-by-step", "rollback-and-recovery-plan", "test-plan-matrix"]),
                FlowAgentStep(3, "role-implementation-lead", "feature-implementation", "implementation", "Apply the data-layer changes in safe slices.", ["implement-in-small-slices", "mandatory-unit-tests", "mandatory-integration-tests"]),
                FlowAgentStep(4, "role-test-validation-lead", "validation-audit", "delivery", "Verify migrations, provider compatibility, and rollback notes.", ["final-audit", "evidence-output-required"]),
            ],
        ),
        FlowDefinition(
            key="playwright-automation-upgrade",
            name="Playwright Automation Upgrade",
            summary="Flow for expanding browser automation, diagnostics, and reliable UI evidence.",
            prompt_types=["testing", "validation", "ui", "implementation"],
            block_keys=["mission-exact-goal", "success-criteria", "current-state-audit", "environment-and-commands", "test-plan-matrix", "mandatory-playwright-tests", "manual-verification-steps", "evidence-output-required", "final-audit"],
            agent_sequence=[
                FlowAgentStep(1, "role-test-validation-lead", "test-strategy-and-automation", "planning", "Audit current automation and define the target test matrix.", ["current-state-audit", "test-plan-matrix", "environment-and-commands"]),
                FlowAgentStep(2, "role-implementation-lead", "feature-implementation", "implementation", "Implement or repair the UI automation.", ["mandatory-playwright-tests"]),
                FlowAgentStep(3, "role-test-validation-lead", "validation-audit", "delivery", "Gather artifacts and close out the browser proof.", ["manual-verification-steps", "evidence-output-required", "final-audit"]),
            ],
        ),
        FlowDefinition(
            key="php-canvas-migration",
            name="PHP Canvas Modernization and Migration",
            summary="Flow for PHP apps adding or rebuilding advanced canvas-based UI with controlled modernization.",
            prompt_types=["architecture", "implementation", "migration", "ui", "validation"],
            block_keys=["mission-exact-goal", "scope-in-scope-items", "current-state-audit", "parity-matrix", "ux-surface-map", "implementation-plan-step-by-step", "implement-in-small-slices", "mandatory-playwright-tests", "final-audit"],
            agent_sequence=[
                FlowAgentStep(1, "role-ui-ux-engineer", "architecture-spec", "architecture", "Map existing PHP surfaces to the target canvas-first flow.", ["current-state-audit", "parity-matrix", "ux-surface-map"]),
                FlowAgentStep(2, "role-implementation-planner", "implementation-plan", "planning", "Sequence reusable primitives, host integration, and tests.", ["implementation-plan-step-by-step", "test-plan-matrix"]),
                FlowAgentStep(3, "role-implementation-lead", "feature-implementation", "implementation", "Implement the migration in stable increments.", ["implement-in-small-slices", "mandatory-playwright-tests"]),
                FlowAgentStep(4, "role-test-validation-lead", "validation-audit", "delivery", "Validate browser behavior and migration completeness.", ["final-audit", "evidence-output-required"]),
            ],
        ),
        FlowDefinition(
            key="embedded-midi-firmware-tuning",
            name="Embedded MIDI and Firmware Tuning",
            summary="Flow for firmware, board constraints, MIDI timing, and telemetry-backed validation.",
            prompt_types=["embedded", "implementation", "testing", "validation", "review"],
            block_keys=["mission-exact-goal", "success-criteria", "required-reading-list", "current-state-audit", "dependency-inventory", "non-negotiable-rules", "safe-ambiguity-handling", "multi-agent-chain", "required-phase-output-format", "risk-register", "implementation-plan-step-by-step", "implement-in-small-slices", "mandatory-unit-tests", "performance-validation-pass", "observability-validation-pass", "manual-verification-steps", "final-audit", "output-completion-summary", "output-remaining-risks-and-next-steps"],
            agent_sequence=[
                FlowAgentStep(1, "role-embedded-midi-engineer", "embedded-firmware-iteration", "architecture", "Audit the signal path, pin usage, and telemetry needs.", ["required-reading-list", "current-state-audit", "risk-register"]),
                FlowAgentStep(2, "role-implementation-planner", "implementation-plan", "planning", "Break the firmware and host-surface changes into safe slices.", ["implementation-plan-step-by-step", "test-plan-matrix"]),
                FlowAgentStep(3, "role-embedded-midi-engineer", "feature-implementation", "implementation", "Apply the firmware and timing changes with observability.", ["implement-in-small-slices", "mandatory-unit-tests", "observability-validation-pass"]),
                FlowAgentStep(4, "role-test-validation-lead", "validation-audit", "delivery", "Validate timing, performance, telemetry, and manual hardware flows.", ["performance-validation-pass", "manual-verification-steps", "final-audit"]),
            ],
        ),
        FlowDefinition(
            key="release-hardening-final-audit",
            name="Release Hardening and Final Audit",
            summary="Flow for final proof, hardening, and release readiness.",
            prompt_types=["validation", "performance", "security", "review"],
            block_keys=["success-criteria", "definition-of-done", "reviewer-findings-first", "architecture-validation-pass", "security-validation-pass", "performance-validation-pass", "observability-validation-pass", "final-audit", "evidence-output-required", "output-commands-executed", "output-remaining-risks-and-next-steps"],
            agent_sequence=[
                FlowAgentStep(1, "role-senior-reviewer", "validation-audit", "verification", "Review the release candidate for missing proof and high-risk flaws.", ["reviewer-findings-first", "architecture-validation-pass", "security-validation-pass", "performance-validation-pass"]),
                FlowAgentStep(2, "role-test-validation-lead", "validation-audit", "delivery", "Collect final evidence and call out residual risk honestly.", ["observability-validation-pass", "final-audit", "evidence-output-required", "output-remaining-risks-and-next-steps"]),
            ],
        ),
    ]


def build_blueprints() -> list[BlueprintDefinition]:
    return [
        BlueprintDefinition("architecture-spec", "Architecture Spec", "architecture", "Creates an implementation-ready architecture, gap analysis, risk model, and recommended next steps.", "Use when the first agent needs to design the solution, clarify boundaries, or map a major migration or feature before coding starts.", "architecture-review-plan-implement-validate", ["role-architecture-lead", "mission-exact-goal", "repo-map-confirmation", "required-reading-list", "architecture-blueprint", "risk-register"]),
        BlueprintDefinition("repository-audit", "Repository Audit", "audit", "Audits the current state of the repository, identifies gaps, and creates a target map.", "Use when the repo is only partially understood, the bundle was created off-repo, or current implementation claims need to be verified before planning.", "audit-plan-refactor-review", ["role-refactor-specialist", "current-state-audit", "file-touch-plan", "gap-analysis", "parity-matrix"]),
        BlueprintDefinition("implementation-plan", "Implementation Plan", "plan", "Converts an approved design or goal into milestones, files, tests, and acceptance checkpoints.", "Use after architecture or audit work to prepare a deterministic plan that another agent can execute.", "architecture-review-plan-implement-validate", ["role-implementation-planner", "implementation-plan-step-by-step", "dependency-ordering", "test-plan-matrix", "definition-of-done"]),
        BlueprintDefinition("feature-implementation", "Feature Implementation", "implementation", "Delivers a feature or enhancement in code with staged verification and handoff.", "Use when the main requirement is to produce working code, tests, and artifacts rather than just planning.", "architecture-review-plan-implement-validate", ["role-implementation-lead", "implement-in-small-slices", "mandatory-unit-tests", "final-audit", "output-completion-summary"]),
        BlueprintDefinition("safe-refactor", "Safe Refactor", "refactor", "Improves structure while preserving behavior and locking regressions down.", "Use when complexity or duplication is the main issue and behavior should stay stable.", "audit-plan-refactor-review", ["role-refactor-specialist", "current-state-audit", "preserve-backward-compatibility", "additive-refactor-first", "regression-proof-required"]),
        BlueprintDefinition("bugfix-with-regression-lock", "Bugfix with Regression Lock", "bugfix", "Fixes a bug and leaves behind targeted regression proof.", "Use when the agent must reproduce, fix, and lock down a concrete defect quickly without widening scope.", "bugfix-regression-proof", ["role-implementation-lead", "success-criteria", "mandatory-unit-tests", "regression-proof-required", "manual-verification-steps"]),
        BlueprintDefinition("senior-code-review", "Senior Code Review", "review", "Produces findings-first review output focused on behavior, risk, and missing evidence.", "Use when a second-pass reviewer should challenge the design, change set, or validation claims.", "release-hardening-final-audit", ["role-senior-reviewer", "reviewer-findings-first", "architecture-validation-pass", "security-validation-pass", "performance-validation-pass"]),
        BlueprintDefinition("test-strategy-and-automation", "Test Strategy and Automation", "testing", "Designs or expands the test matrix, automation approach, and evidence expectations.", "Use when the session is about validation design, Playwright rollout, or closing coverage gaps.", "playwright-automation-upgrade", ["role-test-validation-lead", "test-plan-matrix", "mandatory-playwright-tests", "evidence-output-required"]),
        BlueprintDefinition("validation-audit", "Validation Audit", "validation", "Performs a final proof-oriented audit on architecture, testing, and residual risk.", "Use when code or plans already exist and the main job is to validate, challenge, and close out honestly.", "release-hardening-final-audit", ["role-test-validation-lead", "final-audit", "evidence-output-required", "output-remaining-risks-and-next-steps"]),
        BlueprintDefinition("performance-hardening", "Performance Hardening", "performance", "Optimizes hot paths, runtime cost, or scheduling behavior with explicit proof.", "Use when latency, browser performance, rendering cost, or embedded timing is part of the quality bar.", "release-hardening-final-audit", ["performance-validation-pass", "risk-register", "manual-verification-steps", "final-audit"]),
        BlueprintDefinition("security-hardening", "Security Hardening", "security", "Focuses the session on threats, trust boundaries, secret handling, and hardening proof.", "Use for auth, billing, secrets, client/server trust boundaries, or unsafe configuration changes.", "release-hardening-final-audit", ["protect-secrets-and-sensitive-data", "security-validation-pass", "risk-register", "final-audit"]),
        BlueprintDefinition("ui-ux-delivery", "UI/UX Delivery", "ui", "Designs and delivers user-facing flows with real interaction, accessibility, and browser proof.", "Use when the prompt is centered on pages, components, canvas, or navigation flows.", "ui-canvas-feature-delivery", ["role-ui-ux-engineer", "ux-surface-map", "no-placeholder-ui", "mandatory-playwright-tests", "accessibility-validation-pass"]),
        BlueprintDefinition("embedded-firmware-iteration", "Embedded Firmware Iteration", "embedded", "Designs and delivers firmware or hardware-integrated changes with timing and hardware constraints visible.", "Use when the work touches microcontrollers, sensors, MIDI, serial protocols, or host-side telemetry for hardware flows.", "embedded-midi-firmware-tuning", ["role-embedded-midi-engineer", "risk-register", "mandatory-unit-tests", "performance-validation-pass", "manual-verification-steps"]),
    ]


def build_simulations() -> list[SimulationCase]:
    return [
        SimulationCase("candoitall-branch-aware-prompt-flow", "CanDoItAll Branch-Aware Prompt Flow Visualization", "Add branch-aware prompt-run visualization and lineage-aware validation to the existing .NET/Blazor Prompt Factory and Workbench.", "architecture-review-plan-implement-validate", [".net", "blazor", "efcore", "postgresql", "tailwind", "playwright", "offline-first"], ["session-framing", "mission-scope", "context-discovery", "guardrails", "workflow-orchestration", "architecture-analysis", "planning-checklists", "implementation-execution", "validation-review", "output-handoff", "stack-profiles", "toolbox-snippets"], ["stack-dotnet-solution", "stack-blazor-webapp", "stack-tailwind-css", "stack-efcore", "stack-postgresql", "stack-playwright-mcp", "stack-offline-first-sync", "toolbox-run-unit-tests-docker", "toolbox-run-ui-tests-docker"], ["role-architecture-lead", "role-senior-reviewer", "role-implementation-planner", "role-implementation-lead", "role-test-validation-lead"], ["architecture", "ui", "postgresql", "docker-tests"]),
        SimulationCase("php-canvas-calendar-recurring-events", "PHP Canvas Calendar Recurring Events and Drag/Drop", "Extend a PHP app with canvas-first recurring events, drag/drop, and Outlook-like dense layouts validated in a real browser.", "ui-canvas-feature-delivery", ["php", "html", "javascript", "css", "canvas", "playwright"], ["session-framing", "mission-scope", "context-discovery", "guardrails", "workflow-orchestration", "architecture-analysis", "planning-checklists", "implementation-execution", "validation-review", "output-handoff", "stack-profiles", "toolbox-snippets"], ["stack-php-webapp", "stack-html-js-css", "stack-canvas-html-js", "stack-playwright-mcp", "toolbox-use-playwright-mcp-now", "toolbox-capture-browser-artifacts"], ["role-ui-ux-engineer", "role-implementation-planner", "role-implementation-lead", "role-test-validation-lead"], ["canvas-interaction", "browser-proof", "accessibility"]),
        SimulationCase("safe-refactor-context-assembly", "Safe Refactor of Prompt Factory Context Assembly", "Refactor the CanDoItAll prompt context assembly pipeline for lower coupling while preserving behavior and locking regressions with Docker tests.", "audit-plan-refactor-review", [".net", "blazor", "efcore", "postgresql", "playwright"], ["session-framing", "mission-scope", "context-discovery", "guardrails", "workflow-orchestration", "planning-checklists", "implementation-execution", "validation-review", "output-handoff", "stack-profiles", "toolbox-snippets"], ["stack-dotnet-solution", "stack-blazor-webapp", "stack-efcore", "stack-postgresql", "stack-playwright-mcp", "toolbox-run-unit-tests-docker", "toolbox-run-integration-tests-docker", "toolbox-cache-downloads-mobile-data"], ["role-refactor-specialist", "role-implementation-planner", "role-senior-reviewer"], ["regression-proof", "docker", "backward-compatibility"]),
        SimulationCase("offline-sync-entitlements-feature", "Offline Entitlements and Sync-Ready Account Feature", "Add an offline-first entitlements and sync-ready account workflow spanning API, EF Core, PostgreSQL, and Blazor UI.", "fullstack-offline-feature", [".net", "blazor", "efcore", "postgresql", "offline-first", "playwright"], ["session-framing", "mission-scope", "context-discovery", "guardrails", "workflow-orchestration", "architecture-analysis", "planning-checklists", "implementation-execution", "validation-review", "output-handoff", "stack-profiles", "toolbox-snippets"], ["stack-dotnet-solution", "stack-blazor-webapp", "stack-efcore", "stack-postgresql", "stack-offline-first-sync", "stack-playwright-mcp", "toolbox-cross-db-compat-check", "toolbox-db-migration-dry-run", "toolbox-run-integration-tests-docker"], ["role-architecture-lead", "role-senior-reviewer", "role-implementation-planner", "role-implementation-lead", "role-test-validation-lead"], ["sync", "postgresql", "migration", "offline-proof"]),
        SimulationCase("m5stack-midi-piezo-hit-engine", "M5Stack MIDI Piezo Hit Engine Refinement", "Refine a piezo-first hit engine on M5Stack with better timing, telemetry, and host-side validation.", "embedded-midi-firmware-tuning", ["arduino", "m5stack", "midi", "audio", ".net", "blazor"], ["session-framing", "mission-scope", "context-discovery", "guardrails", "workflow-orchestration", "architecture-analysis", "planning-checklists", "implementation-execution", "validation-review", "output-handoff", "stack-profiles", "toolbox-snippets"], ["stack-arduino-firmware", "stack-m5stack", "stack-midi-audio", "stack-dotnet-solution", "stack-blazor-webapp", "toolbox-generate-fixtures-and-seed-data", "toolbox-create-manual-qa-checklist"], ["role-embedded-midi-engineer", "role-implementation-planner", "role-test-validation-lead"], ["timing", "hardware-constraints", "telemetry", "manual-hardware-checks"]),
    ]


def analyze_prompt_packs() -> dict:
    text_exts = {".md", ".txt", ".json", ".yaml", ".yml", ".csv", ".mmd"}
    heading_counter = Counter()
    keyword_counter = Counter()
    pack_stats: dict[str, dict[str, int]] = defaultdict(lambda: {"text_files": 0, "promptish_files": 0, "docish_files": 0, "checklists": 0, "readmes": 0})
    keywords = ["review", "tests", "checklist", "goal", "validation", "audit", "branch", "workflow", "playwright", "phase", "mission", "architecture", "risk", "acceptance criteria", "constraints", "deliverables", "definition of done", "file-touch", "artifact"]

    def analyze_text(pack_name: str, rel_path: str, text: str) -> None:
        stats = pack_stats[pack_name]
        stats["text_files"] += 1
        rel_lower = rel_path.lower()
        if any(token in rel_lower for token in ["prompt", "starter", "start"]):
            stats["promptish_files"] += 1
        if "checklist" in rel_lower:
            stats["checklists"] += 1
        if "readme" in rel_lower:
            stats["readmes"] += 1
        if any(token in rel_lower for token in ["docs/", "analysis/", "design/", "implementation/", "architecture/", "references/"]):
            stats["docish_files"] += 1

        for line in text.splitlines():
            match = re.match(r"^(#{1,6})\s+(.+?)\s*$", line.strip())
            if match:
                heading = re.sub(r"\s+", " ", match.group(2).strip().lower())
                heading_counter[heading] += 1

        lower = text.lower()
        for keyword in keywords:
            if keyword in lower:
                keyword_counter[keyword] += 1

    for pack_path in sorted(PACKS_ROOT.iterdir(), key=lambda item: item.name.lower()):
        if pack_path.is_dir():
            for file_path in pack_path.rglob("*"):
                if file_path.suffix.lower() not in text_exts or not file_path.is_file():
                    continue
                analyze_text(pack_path.name, str(file_path.relative_to(pack_path)).replace("\\", "/"), file_path.read_text(encoding="utf-8", errors="ignore"))
        elif pack_path.suffix.lower() == ".zip":
            with zipfile.ZipFile(pack_path) as archive:
                for entry in archive.infolist():
                    if entry.is_dir() or Path(entry.filename).suffix.lower() not in text_exts:
                        continue
                    analyze_text(pack_path.name, entry.filename, archive.read(entry.filename).decode("utf-8", errors="ignore"))

    return {
        "total_packs": len(pack_stats),
        "total_text_files": sum(item["text_files"] for item in pack_stats.values()),
        "pack_stats": dict(sorted(pack_stats.items())),
        "top_headings": heading_counter.most_common(12),
        "top_keywords": keyword_counter.most_common(12),
    }


def placeholder_glossary() -> list[dict[str, str]]:
    return [
        {"token": "exact_goal", "meaning": "The exact feature, fix, or artifact the prompt should focus on."},
        {"token": "target_feature_or_problem", "meaning": "Short name of the feature, module, or problem space."},
        {"token": "business_context", "meaning": "Why the work matters for the user or business."},
        {"token": "in_scope_item_1", "meaning": "A concrete requirement or work item that must be covered."},
        {"token": "out_of_scope_item_1", "meaning": "An adjacent area that should be explicitly excluded."},
        {"token": "success_criterion_1", "meaning": "A measurable success condition."},
        {"token": "deliverable_1", "meaning": "A concrete output artifact such as code, docs, tests, or a checklist."},
        {"token": "solution_or_workspace_root", "meaning": "Top-level repo or solution root that should be confirmed."},
        {"token": "primary_projects_or_modules", "meaning": "Main projects, services, modules, or packages that matter."},
        {"token": "tests_and_validation_projects", "meaning": "The test suites or validation paths tied to the change."},
        {"token": "build_command", "meaning": "Canonical build command for the workspace."},
        {"token": "unit_test_command", "meaning": "Canonical unit test command for the workspace."},
        {"token": "integration_test_command", "meaning": "Canonical integration or API test command."},
        {"token": "ui_test_command", "meaning": "Canonical Playwright or UI test command."},
        {"token": "docker_compose_file_or_dockerfile", "meaning": "Docker asset that should be used for isolated test runs."},
    ]


def build_group_summary_rows(groups: list[GroupDefinition], components: list[ComponentDefinition]) -> list[dict]:
    by_group = defaultdict(list)
    for item in components:
        by_group[item.group].append(item)

    rows = []
    for group in sorted(groups, key=lambda item: item.order):
        members = sorted(by_group[group.key], key=lambda item: item.name)
        rows.append({"key": group.key, "name": group.name, "summary": group.summary, "purpose": group.purpose, "uiMode": group.ui_mode, "order": group.order, "componentCount": len(members), "componentNames": [member.name for member in members], "componentKeys": [member.key for member in members]})
    return rows


def validate_simulations(simulations: list[SimulationCase], components: dict[str, ComponentDefinition], flows: dict[str, FlowDefinition]) -> list[dict]:
    results = []
    for case in simulations:
        flow = flows[case.flow_key]
        role_keys = [step.role_component_key for step in flow.agent_sequence]
        selected_keys = list(dict.fromkeys(flow.block_keys + role_keys + case.extra_block_keys))
        selected_components = [components[key] for key in selected_keys]
        covered_groups = sorted({component.group for component in selected_components})
        covered_stack_tags = sorted({tag for component in selected_components for tag in component.stack_tags})
        covered_roles = sorted({step.role_component_key for step in flow.agent_sequence})

        missing_groups = sorted(set(case.required_groups) - set(covered_groups))
        missing_roles = sorted(set(case.expected_roles) - set(covered_roles))
        missing_stack_tags = sorted(set(case.stack_tags) - set(covered_stack_tags))
        validation_flags = {
            "hasArchitectureAgent": "role-architecture-lead" in covered_roles or "role-ui-ux-engineer" in covered_roles or "role-embedded-midi-engineer" in covered_roles or "role-refactor-specialist" in covered_roles,
            "hasReviewer": "role-senior-reviewer" in covered_roles or "role-test-validation-lead" in covered_roles,
            "hasPlanner": "role-implementation-planner" in covered_roles,
            "hasImplementer": "role-implementation-lead" in covered_roles or "role-embedded-midi-engineer" in covered_roles or "role-refactor-specialist" in covered_roles,
            "hasValidator": "role-test-validation-lead" in covered_roles or "role-senior-reviewer" in covered_roles,
        }

        results.append(
            {
                "id": case.id,
                "key": case.key,
                "name": case.name,
                "summary": case.summary,
                "flowKey": case.flow_key,
                "selectedBlockKeys": selected_keys,
                "selectedBlockNames": [component.name for component in selected_components],
                "coveredGroups": covered_groups,
                "coveredStackTags": covered_stack_tags,
                "coveredRoles": covered_roles,
                "requiredGroups": case.required_groups,
                "requiredStackTags": case.stack_tags,
                "requiredRoles": case.expected_roles,
                "validationFocus": case.validation_focus,
                "missingGroups": missing_groups,
                "missingRoles": missing_roles,
                "missingStackTags": missing_stack_tags,
                "flowValidationFlags": validation_flags,
                "passes": not missing_groups and not missing_roles and not missing_stack_tags and all(validation_flags.values()),
            }
        )
    return results


def ensure_directories() -> None:
    for path in [DOCS_ROOT, OUTPUT_ROOT, SPREADSHEET_ROOT, TEMPLATE_ROOT]:
        path.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    def clean(value: str) -> str:
        return value.replace("\n", "<br>").replace("|", "\\|")

    header_line = "| " + " | ".join(headers) + " |"
    divider_line = "| " + " | ".join(["---"] * len(headers)) + " |"
    body_lines = ["| " + " | ".join(clean(str(cell)) for cell in row) + " |" for row in rows]
    return "\n".join([header_line, divider_line, *body_lines])


def write_template_files(components: list[ComponentDefinition]) -> None:
    for component_def in components:
        group_dir = TEMPLATE_ROOT / component_def.group
        group_dir.mkdir(parents=True, exist_ok=True)
        content = dedent(
            f"""
            ---
            key: {component_def.key}
            id: {component_def.id}
            name: {component_def.name}
            group: {component_def.group}
            blockKind: {component_def.block_kind}
            toolboxEligible: {str(component_def.toolbox_eligible).lower()}
            recommended: {str(component_def.recommended).lower()}
            tags: {", ".join(component_def.tags)}
            promptTypes: {", ".join(component_def.prompt_types)}
            blueprints: {", ".join(component_def.blueprints)}
            phases: {", ".join(component_def.phases)}
            stackTags: {", ".join(component_def.stack_tags)}
            templateTokens: {", ".join(component_def.template_tokens)}
            ---

            {component_def.template}
            """
        )
        (group_dir / f"{component_def.key}.md").write_text(content + "\n", encoding="utf-8")


def write_csv(path: Path, components: list[ComponentDefinition]) -> None:
    fieldnames = ["id", "key", "name", "group", "blockKind", "summary", "tags", "promptTypes", "blueprints", "phases", "stackTags", "toolboxEligible", "recommended", "templateTokens", "templatePart"]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for component_def in components:
            writer.writerow({"id": component_def.id, "key": component_def.key, "name": component_def.name, "group": component_def.group, "blockKind": component_def.block_kind, "summary": component_def.summary, "tags": ", ".join(component_def.tags), "promptTypes": ", ".join(component_def.prompt_types), "blueprints": ", ".join(component_def.blueprints), "phases": ", ".join(component_def.phases), "stackTags": ", ".join(component_def.stack_tags), "toolboxEligible": "yes" if component_def.toolbox_eligible else "no", "recommended": "yes" if component_def.recommended else "no", "templateTokens": ", ".join(component_def.template_tokens), "templatePart": component_def.template})


def style_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9E1F2")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    max_widths: dict[int, int] = defaultdict(int)
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            width = min(max((len(line) for line in value.splitlines()), default=0), 90)
            max_widths[cell.column] = max(max_widths[cell.column], width)
    for column_index, width in max_widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = max(12, width + 2)


def write_workbook(workbook_path: Path, components: list[ComponentDefinition], groups: list[GroupDefinition], blueprints: list[BlueprintDefinition], flows: list[FlowDefinition], simulation_results: list[dict], pack_analysis: dict) -> None:
    workbook = Workbook()
    components_sheet = workbook.active
    components_sheet.title = "Components"
    components_sheet.append(["Id", "Key", "Name", "Group", "BlockKind", "Summary", "Tags", "PromptTypes", "Blueprints", "Phases", "StackTags", "ToolboxEligible", "Recommended", "TemplateTokens", "TemplatePart"])
    for component_def in components:
        components_sheet.append([component_def.id, component_def.key, component_def.name, component_def.group, component_def.block_kind, component_def.summary, ", ".join(component_def.tags), ", ".join(component_def.prompt_types), ", ".join(component_def.blueprints), ", ".join(component_def.phases), ", ".join(component_def.stack_tags), "yes" if component_def.toolbox_eligible else "no", "yes" if component_def.recommended else "no", ", ".join(component_def.template_tokens), component_def.template])
    style_sheet(components_sheet)

    groups_sheet = workbook.create_sheet("Groups")
    groups_sheet.append(["Key", "Name", "Order", "UI Mode", "Summary", "Purpose", "ComponentCount"])
    for row in build_group_summary_rows(groups, components):
        groups_sheet.append([row["key"], row["name"], row["order"], row["uiMode"], row["summary"], row["purpose"], row["componentCount"]])
    style_sheet(groups_sheet)

    blueprints_sheet = workbook.create_sheet("Blueprints")
    blueprints_sheet.append(["Id", "Key", "Name", "PromptType", "Summary", "Guidance", "RecommendedFlowKey", "RecommendedBlockKeys"])
    for blueprint in blueprints:
        blueprints_sheet.append([blueprint.id, blueprint.key, blueprint.name, blueprint.prompt_type, blueprint.summary, blueprint.guidance, blueprint.recommended_flow_key, ", ".join(blueprint.recommended_block_keys)])
    style_sheet(blueprints_sheet)

    flows_sheet = workbook.create_sheet("Flows")
    flows_sheet.append(["Id", "Key", "Name", "Summary", "PromptTypes", "BlockKeys", "AgentSequence"])
    for flow in flows:
        sequence = "\n".join(f"{step.order}. {step.role_component_key} -> {step.blueprint_key} ({step.phase}) - {step.goal}" for step in flow.agent_sequence)
        flows_sheet.append([flow.id, flow.key, flow.name, flow.summary, ", ".join(flow.prompt_types), ", ".join(flow.block_keys), sequence])
    style_sheet(flows_sheet)

    simulation_sheet = workbook.create_sheet("Simulations")
    simulation_sheet.append(["Id", "Key", "Name", "FlowKey", "Passes", "RequiredGroups", "CoveredGroups", "RequiredStackTags", "CoveredStackTags", "RequiredRoles", "CoveredRoles", "MissingGroups", "MissingStackTags", "MissingRoles", "ValidationFocus"])
    for result in simulation_results:
        simulation_sheet.append([result["id"], result["key"], result["name"], result["flowKey"], "yes" if result["passes"] else "no", ", ".join(result["requiredGroups"]), ", ".join(result["coveredGroups"]), ", ".join(result["requiredStackTags"]), ", ".join(result["coveredStackTags"]), ", ".join(result["requiredRoles"]), ", ".join(result["coveredRoles"]), ", ".join(result["missingGroups"]), ", ".join(result["missingStackTags"]), ", ".join(result["missingRoles"]), ", ".join(result["validationFocus"])])
    style_sheet(simulation_sheet)

    analysis_sheet = workbook.create_sheet("PackAnalysis")
    analysis_sheet.append(["Metric", "Value"])
    analysis_sheet.append(["Total packs", pack_analysis["total_packs"]])
    analysis_sheet.append(["Total text files", pack_analysis["total_text_files"]])
    analysis_sheet.append([])
    analysis_sheet.append(["Top headings", "Count"])
    for heading, count in pack_analysis["top_headings"]:
        analysis_sheet.append([heading, count])
    analysis_sheet.append([])
    analysis_sheet.append(["Top keywords", "Count"])
    for keyword, count in pack_analysis["top_keywords"]:
        analysis_sheet.append([keyword, count])
    style_sheet(analysis_sheet)

    placeholders_sheet = workbook.create_sheet("Placeholders")
    placeholders_sheet.append(["Token", "Meaning"])
    for item in placeholder_glossary():
        placeholders_sheet.append([item["token"], item["meaning"]])
    style_sheet(placeholders_sheet)
    workbook.save(workbook_path)


def write_docs(groups: list[GroupDefinition], components: list[ComponentDefinition], blueprints: list[BlueprintDefinition], flows: list[FlowDefinition], simulations: list[SimulationCase], simulation_results: list[dict], pack_analysis: dict) -> None:
    group_rows = build_group_summary_rows(groups, components)
    components_by_group = defaultdict(list)
    for component_def in components:
        components_by_group[component_def.group].append(component_def)

    recommended_count = sum(1 for component_def in components if component_def.recommended)
    toolbox_count = sum(1 for component_def in components if component_def.toolbox_eligible)
    stack_count = sum(1 for component_def in components if component_def.group == "stack-profiles")
    pass_count = sum(1 for result in simulation_results if result["passes"])

    readme = clean_markdown_doc(
        f"""
        # Prompt Wizard Library

        This library was generated from a direct analysis of the prompt packs in `inputs/prompts packs` and expanded with additional agentic-coding patterns needed for CanDoItAll's prompt wizard and manager.

        ## What is included
        - {len(components)} reusable prompt components
        - {len(groups)} component groups
        - {len(blueprints)} blueprint types
        - {len(flows)} flow templates
        - {len(simulations)} simulation cases with coverage validation
        - import-friendly JSON seed files aligned to `CanDoItAll.Modules.Factory`
        - markdown snippet files for each component
        - an Excel catalog at `output/spreadsheet/prompt-component-library.xlsx`

        ## Counts by category
        {markdown_table(["Group", "Components", "UI Mode", "Purpose"], [[row["name"], str(row["componentCount"]), row["uiMode"], row["purpose"]] for row in group_rows])}

        ## Coverage status
        {pass_count} of {len(simulations)} simulation cases pass the required group, role, and stack coverage checks.
        """
    )
    (DOCS_ROOT / "README.md").write_text(readme + "\n", encoding="utf-8")

    analysis_doc = clean_markdown_doc(
        f"""
        # Analysis and Best Practices

        The prompt-pack scan covered {pack_analysis["total_packs"]} packs and {pack_analysis["total_text_files"]} text-like files across folders and zip bundles.

        ## Most repeated headings
        {markdown_table(["Heading", "Count"], [[heading, str(count)] for heading, count in pack_analysis["top_headings"]])}

        ## Most repeated keywords
        {markdown_table(["Keyword", "Count"], [[keyword, str(count)] for keyword, count in pack_analysis["top_keywords"]])}

        ## Conclusions
        1. The best packs behave like workflows, not standalone prompts.
        2. They force repo inspection before design or coding.
        3. They define hard rules and scope boundaries explicitly.
        4. They split work into milestones or gated phases.
        5. They require tests, validation, and evidence after each significant step.
        6. They keep continuity through checklists, decisions logs, and next-prompt pointers.
        """
    )
    (DOCS_ROOT / "01-analysis-and-best-practices.md").write_text(analysis_doc + "\n", encoding="utf-8")

    architecture_doc = clean_markdown_doc(
        f"""
        # Component Architecture

        ## Composition order
        1. role and mission
        2. context loading
        3. guardrails
        4. architecture or planning
        5. implementation or validation
        6. stack profiles
        7. toolbox snippets
        8. output and handoff

        ## Relationship to the current app model
        - `PromptBlockDefinition`: atomic reusable block
        - `PromptBlueprint`: prompt type plus recommended flow
        - `PromptFlowTemplate`: ordered workflow and agent sequence
        - `PromptRun` and `PromptRunNode`: execution lineage
        - `ValidationRun`: quality and review pass

        ## Groups
        {markdown_table(["Order", "Group", "UI Mode", "Component Count", "Purpose"], [[str(row["order"]), row["name"], row["uiMode"], str(row["componentCount"]), row["purpose"]] for row in group_rows])}

        ## Blueprints
        {markdown_table(["Blueprint", "Prompt Type", "Recommended Flow", "Summary"], [[item.name, item.prompt_type, item.recommended_flow_key, item.summary] for item in blueprints])}

        ## Placeholder glossary
        {markdown_table(["Token", "Meaning"], [[item["token"], item["meaning"]] for item in placeholder_glossary()])}
        """
    )
    (DOCS_ROOT / "02-component-architecture.md").write_text(architecture_doc + "\n", encoding="utf-8")

    group_sections = []
    for group in sorted(groups, key=lambda item: item.order):
        members = sorted(components_by_group[group.key], key=lambda item: item.name)
        group_sections.append(
            "\n".join(
                [
                    f"## {group.name}",
                    group.summary,
                    "",
                    markdown_table(["Component", "Key", "BlockKind", "Toolbox", "Recommended"], [[member.name, member.key, member.block_kind, "yes" if member.toolbox_eligible else "no", "yes" if member.recommended else "no"] for member in members]),
                    "",
                ]
            )
        )

    plan_doc = clean_markdown_doc(
        f"""
        # Groups, Checklists, and Creation Plan

        - Recommended core blocks: {recommended_count}
        - Toolbox-ready blocks: {toolbox_count}
        - Stack profile blocks: {stack_count}

        ## Executed creation plan
        1. Analyze prompt packs for recurring structure and proof obligations.
        2. Normalize the library into atomic groups, blueprints, and flows.
        3. Add supported stack adapters and quick-insert toolbox snippets.
        4. Validate the library against multi-agent simulation cases.
        5. Export JSON, markdown snippets, CSV, and XLSX artifacts.

        {"\n".join(group_sections)}
        """
    )
    (DOCS_ROOT / "03-groups-checklists-and-plan.md").write_text(plan_doc + "\n", encoding="utf-8")

    simulation_rows = [[result["name"], result["flowKey"], "PASS" if result["passes"] else "FAIL", ", ".join(result["validationFocus"]), ", ".join(result["missingGroups"]) or "-", ", ".join(result["missingStackTags"]) or "-", ", ".join(result["missingRoles"]) or "-"] for result in simulation_results]
    simulation_doc = clean_markdown_doc(
        f"""
        # Simulation Validation

        {markdown_table(["Case", "Flow", "Result", "Validation Focus", "Missing Groups", "Missing Stack Tags", "Missing Roles"], simulation_rows)}

        ## Cases
        {"\n\n".join([f"## {case.name}\n- Summary: {case.summary}\n- Flow: {case.flow_key}\n- Required stack tags: {', '.join(case.stack_tags)}\n- Expected roles: {', '.join(case.expected_roles)}\n- Extra block inserts: {', '.join(case.extra_block_keys)}" for case in simulations])}
        """
    )
    (DOCS_ROOT / "04-simulation-validation.md").write_text(simulation_doc + "\n", encoding="utf-8")


def write_output_readme(components: list[ComponentDefinition], flows: list[FlowDefinition], blueprints: list[BlueprintDefinition], simulation_results: list[dict]) -> None:
    pass_count = sum(1 for item in simulation_results if item["passes"])
    readme = clean_markdown_doc(
        f"""
        # Prompt Library Outputs

        - Components: {len(components)}
        - Flows: {len(flows)}
        - Blueprints: {len(blueprints)}
        - Simulation cases passing: {pass_count} / {len(simulation_results)}

        Files:
        - `prompt-component-library.json`
        - `factory-prompt-blocks.seed.json`
        - `factory-prompt-flow-templates.seed.json`
        - `factory-prompt-blueprints.seed.json`
        - `group-catalog.json`
        - `simulation-results.json`
        - `prompt-component-library.csv`
        - `templates/`
        """
    )
    (OUTPUT_ROOT / "README.md").write_text(readme + "\n", encoding="utf-8")


def main() -> None:
    ensure_directories()
    groups = GROUPS
    components = sorted(build_components(), key=lambda item: (next(group.order for group in groups if group.key == item.group), item.name))
    flows = build_flows()
    blueprints = build_blueprints()

    component_lookup = {component_def.key: component_def for component_def in components}
    flow_lookup = {flow.key: flow for flow in flows}
    for flow in flows:
        for key in flow.block_keys:
            if key not in component_lookup:
                raise KeyError(f"Flow {flow.key} references missing component {key}")
        for step in flow.agent_sequence:
            if step.role_component_key not in component_lookup:
                raise KeyError(f"Flow {flow.key} references missing role component {step.role_component_key}")

    simulations = build_simulations()
    simulation_results = validate_simulations(simulations, component_lookup, flow_lookup)
    if not all(item["passes"] for item in simulation_results):
        failed = [item["key"] for item in simulation_results if not item["passes"]]
        raise ValueError(f"Simulation coverage failed: {failed}")

    pack_analysis = analyze_prompt_packs()
    group_catalog = build_group_summary_rows(groups, components)
    blocks_seed = [component_def.to_factory_seed() for component_def in components]
    flow_seed = [flow.to_factory_seed(component_lookup) for flow in flows]
    blueprint_seed = [item.to_factory_seed(flow_lookup) for item in blueprints]
    manifest = {"version": 1, "generatedBy": "tools/prompt_library/build_prompt_component_library.py", "componentCount": len(components), "flowCount": len(flows), "blueprintCount": len(blueprints), "simulationCount": len(simulations), "recommendedComponentCount": sum(1 for item in components if item.recommended), "toolboxComponentCount": sum(1 for item in components if item.toolbox_eligible)}

    write_json(OUTPUT_ROOT / "manifest.json", manifest)
    write_json(OUTPUT_ROOT / "prompt-component-library.json", blocks_seed)
    write_json(OUTPUT_ROOT / "factory-prompt-blocks.seed.json", blocks_seed)
    write_json(OUTPUT_ROOT / "factory-prompt-flow-templates.seed.json", flow_seed)
    write_json(OUTPUT_ROOT / "factory-prompt-blueprints.seed.json", blueprint_seed)
    write_json(OUTPUT_ROOT / "group-catalog.json", group_catalog)
    write_json(OUTPUT_ROOT / "simulation-results.json", simulation_results)
    write_json(OUTPUT_ROOT / "prompt-flow-library.json", [asdict(flow) for flow in flows])
    write_json(OUTPUT_ROOT / "blueprint-library.json", [asdict(item) for item in blueprints])
    write_json(OUTPUT_ROOT / "pack-analysis-summary.json", pack_analysis)
    write_template_files(components)
    write_csv(OUTPUT_ROOT / "prompt-component-library.csv", components)
    write_workbook(SPREADSHEET_ROOT / "prompt-component-library.xlsx", components, groups, blueprints, flows, simulation_results, pack_analysis)
    write_docs(groups, components, blueprints, flows, simulations, simulation_results, pack_analysis)
    write_output_readme(components, flows, blueprints, simulation_results)
    print(json.dumps({"components": len(components), "flows": len(flows), "blueprints": len(blueprints), "simulations": len(simulations), "workbook": str(SPREADSHEET_ROOT / "prompt-component-library.xlsx"), "docsRoot": str(DOCS_ROOT), "outputRoot": str(OUTPUT_ROOT)}, indent=2))


if __name__ == "__main__":
    main()
